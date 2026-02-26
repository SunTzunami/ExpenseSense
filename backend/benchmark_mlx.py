"""
benchmark_mlx.py  –  MLX multi-model latency × accuracy benchmark
                     for the Router → Specialist → Summarizer pipeline.

Fully simulates the real app flow:
  Router (router_prompt.txt) → Specialist (tool_prompts.py) → Summarizer (summary_prompt.txt)

Optimized for Apple Silicon (M1 Pro, 8 GB RAM):
  - Incremental CSV checkpoint after every rep (crash-safe)
  - Explicit model unloading between models to reclaim memory
  - gc.collect() + mx.metal.clear_cache() after each model run
  - All heavy imports (matplotlib, openpyxl) deferred until after inference

Run from your project root (javascript_app/backend/) or repo root.

Usage:
    python benchmark_mlx.py                   # default reps
    python benchmark_mlx.py quick             # 1 rep smoke-test
    python benchmark_mlx.py --reps 5
    python benchmark_mlx.py --reps 5 --output my_results.xlsx --plots figures/
"""

from __future__ import annotations

import argparse
import csv
import gc
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from statistics import mean, median
from typing import Any, Optional

import numpy as np

# ── Logging: configure root first, then silence mlx_lm ───────────────────────
logging.basicConfig(level=logging.ERROR)
logging.getLogger("mlx_lm").setLevel(logging.ERROR)

# ── Resolve backend root ──────────────────────────────────────────────────────
_CANDIDATE_ROOTS = [
    os.getcwd(),
    os.path.join(os.getcwd(), "javascript_app", "backend"),
    os.path.join(os.getcwd(), "backend"),
]
_BACKEND_ROOT = os.getcwd()
for _root in _CANDIDATE_ROOTS:
    if os.path.isdir(os.path.join(_root, "utils")):
        _BACKEND_ROOT = _root
        if _root not in sys.path:
            sys.path.insert(0, _root)
        break

from utils.mlx_utils import mlx_model
from utils.tool_prompts import get_tool_prompt

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────

MLX_MODELS: list[str] = [
    "Qwen/Qwen3-0.6B-MLX-4bit",
    "mlx-community/Qwen3-0.6B-8bit",
    "lmstudio-community/Qwen3-0.6B-MLX-bf16",
    "mlx-community/granite-4.0-h-micro-4bit",
    "lmstudio-community/Qwen3-1.7B-MLX-4bit",
    "lmstudio-community/Qwen3-1.7B-MLX-8bit",
    # "mlx-community/granite-4.0-h-micro-8bit",
    # Add more MLX model paths / LM Studio IDs here
]

REPETITIONS_PER_Q: int = 3
WARMUP_REPS: int = 1

# Accuracy grade weights (must sum to 1.0)
GRADE_WEIGHTS: dict[str, float] = {
    "router": 0.30,
    "fn":     0.50,
    "param":  0.20,
}

# ──────────────────────────────────────────────────────────────────────────────
# PROMPT LOADERS
# ──────────────────────────────────────────────────────────────────────────────
_PROMPT_SEARCH_ROOTS: list[str] = [
    os.path.join(_BACKEND_ROOT, "utils", "prompts"),
    os.path.join(os.getcwd(), "javascript_app", "backend", "utils", "prompts"),
    os.path.join(os.getcwd(), "utils", "prompts"),
]


def load_prompt_template(filename: str) -> str:
    for base in _PROMPT_SEARCH_ROOTS:
        path = os.path.join(base, filename)
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()
                print(f"  [prompt] Loaded '{filename}' <- {path}")
                return content
            except OSError as e:
                print(f"  [prompt] Error reading {path}: {e}")
    print(f"  [prompt] WARNING: '{filename}' not found in {_PROMPT_SEARCH_ROOTS}")
    return ""


ROUTER_PROMPT  = load_prompt_template("router_prompt.txt")
SUMMARY_PROMPT = load_prompt_template("summary_prompt.txt")

# ──────────────────────────────────────────────────────────────────────────────
# METADATA
# ──────────────────────────────────────────────────────────────────────────────
def _load_metadata():
    try:
        categories_path = os.path.join(_BACKEND_ROOT, "..", "src", "utils", "categories.json")
        with open(categories_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            major_cats = list(data.get("CATEGORY_COLORS", {}).keys())
            sub_cats = list(data.get("CATEGORY_MAPPING", {}).keys())
            
            all_cats = sorted(set(major_cats + sub_cats))
            md = "\n\n### CATEGORIES (available options for 'category' argument):\n"
            for c in all_cats:
                md += f"- {c}\n"
            return md
    except Exception as e:
        print(f"Error loading categories for benchmark: {e}")
        return "\n\n### CATEGORIES (available options for 'category' argument):\n- None\n"

METADATA = _load_metadata()

# ──────────────────────────────────────────────────────────────────────────────
# TEST SUITE  (20 questions)
# ──────────────────────────────────────────────────────────────────────────────
TEST_CASES: list[dict[str, Any]] = [
    # ── Comparison ──────────────────────────────────────────────────────────
    {
        "id": "TC01", "category": "Comparison",
        "q": "compare spending on futsal 2024 vs 2025",
        "tool": "plot_comparison_bars",
        "params": ["category=", "y1=", "y2="],
        "fn_pattern": r"plot_comparison_bars\(df.*category=['\"]futsal game['\"].*y1=(2024|2025).*y2=(2024|2025)",
    },
    {
        "id": "TC02", "category": "Comparison",
        "q": "compare spending on grocery 2025 jan vs 2026 jan",
        "tool": "plot_comparison_bars",
        "params": ["category=", "y1=", "m1=", "y2=", "m2="],
        "fn_pattern": r"plot_comparison_bars\(df.*category=['\"]grocery['\"].*y1=2025.*m1=1.*y2=2026.*m2=1",
    },
    {
        "id": "TC03", "category": "Comparison",
        "q": "compare spending on dining 2024 vs 2025",
        "tool": "plot_comparison_bars",
        "params": ["category=", "y1=", "y2="],
        "fn_pattern": r"plot_comparison_bars\(df.*category=['\"]dining['\"].*y1=(2024|2025).*y2=(2024|2025)",
    },
    {
        "id": "TC04", "category": "Comparison",
        "q": "Compare phone bill Dec 2024 vs Dec 2025",
        "tool": "plot_comparison_bars",
        "params": ["category=", "y1=", "m1=", "y2=", "m2="],
        "fn_pattern": r"plot_comparison_bars\(df.*category=['\"]phone bill['\"].*y1=2024.*m1=12.*y2=2025.*m2=12",
    },
    # ── Time-Series ──────────────────────────────────────────────────────────
    {
        "id": "TC05", "category": "Time-Series",
        "q": "make a plot of how much ive spent on snacks for the past 6 months",
        "tool": "plot_time_series",
        "params": ["category=", "months="],
        "fn_pattern": r"plot_time_series\(df.*category=['\"]snacks['\"].*months=6",
    },
    {
        "id": "TC06", "category": "Time-Series",
        "q": "make a plot to show spend on education for last 12 months",
        "tool": "plot_time_series",
        "params": ["category=", "months="],
        "fn_pattern": r"plot_time_series\(df.*category=['\"]education['\"].*months=12",
    },
    {
        "id": "TC07", "category": "Time-Series",
        "q": "Spending trend for fitness this year",
        "tool": "plot_time_series",
        "params": ["category="],
        "fn_pattern": r"plot_time_series\(df.*category=['\"]Fitness['\"]",
    },
    {
        "id": "TC08", "category": "Time-Series",
        "q": "show me my futsal expenses over time in 2024",
        "tool": "plot_time_series",
        "params": ["category=", "year="],
        "fn_pattern": r"plot_time_series\(df.*category=['\"]futsal game['\"].*year=2024",
    },
    # ── Totals ───────────────────────────────────────────────────────────────
    {
        "id": "TC09", "category": "Totals",
        "q": "how much have i spent on gym in 2024?",
        "tool": "calculate_total",
        "params": ["category=", "year="],
        "fn_pattern": r"calculate_total\(df.*category=['\"]gym['\"].*year=2024",
    },
    {
        "id": "TC10", "category": "Totals",
        "q": "how much did i spend at cafes in 2025?",
        "tool": "calculate_total",
        "params": ["category=", "year="],
        "fn_pattern": r"calculate_total\(df.*category=['\"]cafe['\"].*year=2025",
    },
    {
        "id": "TC11", "category": "Totals",
        "q": "How much did I spend on eating from combinis in 2025?",
        "tool": "calculate_total",
        "params": ["category=", "year="],
        "fn_pattern": r"calculate_total\(df.*category=['\"]combini meal['\"].*year=2025",
    },
    {
        "id": "TC12", "category": "Totals",
        "q": "Total spending from 2023 to 2025",
        "tool": "calculate_total",
        "params": ["start_year=", "end_year="],
        "fn_pattern": r"calculate_total\(df.*start_year=2023.*end_year=2025",
    },
    # ── Distribution ─────────────────────────────────────────────────────────
    {
        "id": "TC13", "category": "Distribution",
        "q": "Show me a pie chart of my spending in 2024",
        "tool": "plot_distribution",
        "params": ["year="],
        "fn_pattern": r"plot_distribution\(df.*year=2024",
    },
    {
        "id": "TC14", "category": "Distribution",
        "q": "Breakdown of housing expenses in 2025",
        "tool": "plot_distribution",
        "params": ["category=", "year="],
        "fn_pattern": r"plot_distribution\(df.*category=['\"]Housing and Utilities['\"].*year=2025",
    },
    {
        "id": "TC15", "category": "Distribution",
        "q": "Distribution of snacks spending in 2025",
        "tool": "plot_distribution",
        "params": ["category=", "year="],
        "fn_pattern": r"plot_distribution\(df.*category=['\"]snacks['\"].*year=2025",
    },
    # ── Statistics ───────────────────────────────────────────────────────────
    {
        "id": "TC16", "category": "Statistics",
        "q": "Average dining expense in 2024?",
        "tool": "calculate_statistics",
        "params": ["category=", "y1="],
        "fn_pattern": r"calculate_statistics\(df.*category=['\"]dining['\"].*y1=2024",
    },
    {
        "id": "TC17", "category": "Statistics",
        "q": "Is there a significant difference in my food spending between 2024 and 2025?",
        "tool": "calculate_statistics",
        "params": ["category=", "y1=", "y2=", "compare="],
        "fn_pattern": r"calculate_statistics\(df.*category=['\"]Food['\"].*y1=(2024|2025).*y2=(2024|2025).*compare=True",
    },
    {
        "id": "TC18", "category": "Statistics",
        "q": "Average monthly spending on electricity in 2024",
        "tool": "calculate_statistics",
        "params": ["category=", "y1="],
        "fn_pattern": r"calculate_statistics\(df.*category=['\"]electricity bill['\"].*y1=2024",
    },
    # ── Top Expenses ─────────────────────────────────────────────────────────
    {
        "id": "TC19", "category": "Top Expenses",
        "q": "What were my biggest expenses in 2025?",
        "tool": "get_top_expenses",
        "params": ["year="],
        "fn_pattern": r"get_top_expenses\(df.*year=2025",
    },
    {
        "id": "TC20", "category": "Top Expenses",
        "q": "Show top 5 grocery purchases in 2025",
        "tool": "get_top_expenses",
        "params": ["n=", "category=", "year="],
        "fn_pattern": r"get_top_expenses\(df.*n=5.*category=['\"]grocery['\"].*year=2025",
    },
]

# Tools that trigger the summarizer stage (exact set, not substring match)
_SUMMARY_TOOLS: frozenset[str] = frozenset({"calculate_statistics", "get_top_expenses"})

# ──────────────────────────────────────────────────────────────────────────────
# INFERENCE
# ──────────────────────────────────────────────────────────────────────────────

def generate_mlx(
    model_id: str,
    messages: list[dict[str, str]],
) -> tuple[str, float, Optional[str]]:
    """Call the MLX model and return (response, elapsed_s, error_or_None)."""
    try:
        t0 = time.perf_counter()
        response = mlx_model.chat(model_id, messages, temperature=0.0)
        return response.strip(), time.perf_counter() - t0, None
    except Exception as exc:
        return "", 0.0, str(exc)


def extract_tool_name(raw_output: str) -> str:
    """Mirror the exact extraction logic used in main.py analyze_stream()."""
    if not raw_output:
        return "NONE"
    return (
        raw_output.split()[0]
        .replace("`", "")
        .replace("'", "")
        .replace('"', "")
        .strip()
    )


# ──────────────────────────────────────────────────────────────────────────────
# ACCURACY GRADING
# ──────────────────────────────────────────────────────────────────────────────

def grade(tc: dict[str, Any], router_tool: str, spec_out: str) -> dict[str, Any]:
    """
    Grade one (router_tool, spec_out) pair against a test-case.

    Router acc  (30%) : exact tool name match
    Fn acc      (50%) : spec output matches expected regex
    Param acc   (20%) : all required params present as 'name=' substrings
    Composite        : weighted sum of the three components
    """
    router_ok = int(router_tool.strip().lower() == tc["tool"].lower())
    fn_ok     = int(bool(re.search(tc["fn_pattern"], spec_out or "", re.IGNORECASE | re.DOTALL)))

    # Each entry in tc["params"] is already 'name=' so we check substring presence.
    # We do this case-insensitively because `llm_input_validation.py` already
    # fuzzy-matches and corrects case on the backend.
    param_ok = int(all(p.lower() in (spec_out or "").lower() for p in tc["params"])) if tc["params"] else 1

    composite = round(
        GRADE_WEIGHTS["router"] * router_ok
        + GRADE_WEIGHTS["fn"]     * fn_ok
        + GRADE_WEIGHTS["param"]  * param_ok,
        4,
    )
    return {
        "router_ok": router_ok,
        "fn_ok":     fn_ok,
        "param_ok":  param_ok,
        "composite": composite,
    }


# ──────────────────────────────────────────────────────────────────────────────
# CHECKPOINT (incremental CSV – crash-safe)
# ──────────────────────────────────────────────────────────────────────────────
_CSV_FIELDNAMES = [
    "Run_Timestamp", "Model", "Model_Full", "TC_ID", "TC_Category",
    "Repetition", "Question", "Expected_Tool", "Detected_Tool",
    "Router_Correct", "Router_Time_s", "Router_Output",
    "Fn_Correct", "Specialist_Time_s", "Specialist_Output",
    "Param_Correct", "Summary_Time_s", "Summary_Output",
    "Composite_Acc", "Total_Time_s",
]


def open_checkpoint(path: str) -> tuple[csv.DictWriter, Any]:
    """
    Open (or create) the incremental CSV checkpoint file.
    Returns (writer, file_handle).  Caller must close the handle.
    """
    exists = os.path.exists(path)
    fh = open(path, "a", newline="", encoding="utf-8")
    writer = csv.DictWriter(fh, fieldnames=_CSV_FIELDNAMES)
    if not exists:
        writer.writeheader()
        fh.flush()
    return writer, fh


def append_checkpoint(writer: csv.DictWriter, fh: Any, row: dict[str, Any]) -> None:
    writer.writerow(row)
    fh.flush()


# ──────────────────────────────────────────────────────────────────────────────
# BENCHMARK LOOP
# ──────────────────────────────────────────────────────────────────────────────

def run_benchmark(
    model_id: str,
    reps: int,
    ckpt_writer: csv.DictWriter,
    ckpt_fh: Any,
) -> list[dict[str, Any]]:
    """
    Run the full pipeline benchmark for one model.
    Writes each non-warmup rep to the CSV checkpoint immediately.
    Returns per-question aggregated stats for console printing.
    """
    short_name   = model_id.split("/")[-1]
    current_date = datetime.now().strftime("%Y-%m-%d")
    run_ts       = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"\n{'='*70}")
    print(f"  Model : {short_name}")
    print(f"  Full  : {model_id}")
    print(f"  Reps  : {reps}  ×  {len(TEST_CASES)} questions  =  {reps * len(TEST_CASES)} calls")
    print(f"{'='*70}")

    per_q_stats: list[dict[str, Any]] = []

    for tc in TEST_CASES:
        print(f"\n  [{tc['id']}] {tc['q']}")

        q_stats: dict[str, Any] = {
            "tc_id":            tc["id"],
            "tc_category":      tc["category"],
            "question":         tc["q"],
            "expected_tool":    tc["tool"],
            "router_times":     [],
            "specialist_times": [],
            "summary_times":    [],
            "total_times":      [],
            "router_ok_sum":    0,
            "fn_ok_sum":        0,
            "param_ok_sum":     0,
            "composite_sum":    0.0,
            "reps":             reps,
        }

        for rep in range(reps + WARMUP_REPS):
            is_warmup = rep < WARMUP_REPS
            # FIX 1: single, unambiguous label — warmups shown as W1, W2…;
            # real reps as 1/N, 2/N, … with no duplicate print line.
            rep_label = f"W{rep + 1}" if is_warmup else f"{rep - WARMUP_REPS + 1}/{reps}"

            # ── Stage 1: Router ───────────────────────────────────────────
            router_out, r_time, r_err = generate_mlx(model_id, [
                {"role": "system", "content": ROUTER_PROMPT},
                {"role": "user",   "content": tc["q"]},
            ])
            tool_name = "ERROR" if r_err else extract_tool_name(router_out)

            # ── Stage 2: Specialist ───────────────────────────────────────
            tool_prompt = get_tool_prompt(tool_name)
            if not tool_prompt:
                tool_name   = "calculate_total"
                tool_prompt = get_tool_prompt("calculate_total")

            system_prompt = tool_prompt.format(
                metadata=METADATA,
                current_date=current_date,
                function_definition=tool_prompt,
            )

            spec_out, s_time, s_err = generate_mlx(model_id, [
                {"role": "system", "content": system_prompt},
                {"role": "user",   "content": tc["q"]},
            ])

            # ── Stage 3: Summarizer (only for qualifying tools) ───────────
            sum_time = 0.0
            sum_out  = ""
            if tool_name in _SUMMARY_TOOLS and not s_err and SUMMARY_PROMPT:
                mock_result = "¥42,000 (n=10, avg ¥4,200)"
                sum_out, sum_time, _ = generate_mlx(model_id, [
                    {"role": "system", "content": SUMMARY_PROMPT},
                    {"role": "user",   "content": (
                        f"User Question: {tc['q']}\n"
                        f"Analysis Result: {mock_result}"
                    )},
                ])

            # ── Grade ─────────────────────────────────────────────────────
            grades     = grade(tc, tool_name, spec_out)
            total_time = r_time + s_time + sum_time

            status = (
                "✓" if grades["composite"] == 1.0
                else ("~" if grades["composite"] > 0 else "✗")
            )

            # FIX 1: single print per rep, warmup clearly labelled
            warmup_tag = "[WARMUP] " if is_warmup else ""
            print(
                f"    Rep {rep_label}  [{status}]  {warmup_tag}"
                f"total={total_time:.2f}s  "
                f"(R={r_time:.2f}s  S={s_time:.2f}s  Sum={sum_time:.2f}s)  "
                f"acc={grades['composite']:.2f}  "
                f"router={'OK' if grades['router_ok'] else 'FAIL'}({tool_name})  "
                f"fn={'OK' if grades['fn_ok'] else 'FAIL'}  "
                f"params={'OK' if grades['param_ok'] else 'FAIL'}"
            )

            # ── Accumulate + checkpoint (warmup excluded from both) ───────
            if not is_warmup:
                q_stats["router_times"].append(r_time)
                q_stats["specialist_times"].append(s_time)
                q_stats["summary_times"].append(sum_time)
                q_stats["total_times"].append(total_time)
                q_stats["router_ok_sum"]  += grades["router_ok"]
                q_stats["fn_ok_sum"]      += grades["fn_ok"]
                q_stats["param_ok_sum"]   += grades["param_ok"]
                q_stats["composite_sum"]  += grades["composite"]

                # FIX 2: checkpoint only written for real reps, never warmup,
                # so the CSV is clean and needs no post-hoc filtering.
                append_checkpoint(ckpt_writer, ckpt_fh, {
                    "Run_Timestamp":     run_ts,
                    "Model":             short_name,
                    "Model_Full":        model_id,
                    "TC_ID":             tc["id"],
                    "TC_Category":       tc["category"],
                    "Repetition":        rep - WARMUP_REPS + 1,
                    "Question":          tc["q"],
                    "Expected_Tool":     tc["tool"],
                    "Detected_Tool":     tool_name,
                    "Router_Correct":    grades["router_ok"],
                    "Router_Time_s":     round(r_time, 3),
                    "Router_Output":     (router_out or "")[:300],
                    "Fn_Correct":        grades["fn_ok"],
                    "Specialist_Time_s": round(s_time, 3),
                    "Specialist_Output": (spec_out or "")[:500],
                    "Param_Correct":     grades["param_ok"],
                    "Summary_Time_s":    round(sum_time, 3),
                    "Summary_Output":    (sum_out or "")[:300],
                    "Composite_Acc":     grades["composite"],
                    "Total_Time_s":      round(total_time, 3),
                })

        per_q_stats.append(q_stats)

    return per_q_stats


# ──────────────────────────────────────────────────────────────────────────────
# MEMORY MANAGEMENT  (critical on 8 GB M1 Pro)
# ──────────────────────────────────────────────────────────────────────────────

def free_model_memory() -> None:
    """
    Best-effort memory reclamation between models.
    mlx_model caches the model internally; we evict it via the public API if
    available, then fall back to gc + Metal cache clear.
    """
    try:
        if hasattr(mlx_model, "unload"):
            mlx_model.unload()
        elif hasattr(mlx_model, "evict"):
            mlx_model.evict()
    except Exception:
        pass

    gc.collect()

    try:
        import mlx.core as mx
        mx.metal.clear_cache()
    except Exception:
        pass


# ──────────────────────────────────────────────────────────────────────────────
# CONSOLE PRINT
# ──────────────────────────────────────────────────────────────────────────────

def print_results(model_id: str, results: list[dict[str, Any]]) -> None:
    short = model_id.split("/")[-1]
    W = 155
    print(f"\n{'='*W}")
    print(f"  {short}")
    print(f"{'='*W}")
    print(
        f"{'TC':<6} {'Category':<15} {'Question':<50} | "
        f"{'R(s)':>7} {'Spec(s)':>8} {'Sum(s)':>7} {'Tot(s)':>7} | "
        f"{'R%':>5} {'Fn%':>5} {'P%':>5} {'Acc%':>6}"
    )
    print("-" * W)

    all_times:     list[float] = []
    all_composite: list[float] = []

    for res in results:
        reps    = res["reps"]
        avg_r   = mean(res["router_times"])
        avg_s   = mean(res["specialist_times"])
        avg_sum = mean(res["summary_times"])
        avg_t   = mean(res["total_times"])
        r_pct   = res["router_ok_sum"]  / reps * 100
        fn_pct  = res["fn_ok_sum"]      / reps * 100
        p_pct   = res["param_ok_sum"]   / reps * 100
        c_pct   = res["composite_sum"]  / reps * 100
        q_disp  = res["question"][:47] + "..." if len(res["question"]) > 50 else res["question"]

        all_times.extend(res["total_times"])
        all_composite.extend(
            [res["composite_sum"] / reps] * len(res["total_times"])
        )

        print(
            f"{res['tc_id']:<6} {res['tc_category']:<15} {q_disp:<50} | "
            f"{avg_r:>6.2f}s {avg_s:>7.2f}s {avg_sum:>6.2f}s {avg_t:>6.2f}s | "
            f"{r_pct:>4.0f}% {fn_pct:>4.0f}% {p_pct:>4.0f}% {c_pct:>5.0f}%"
        )

    print("-" * W)
    overall_r   = mean([t for r in results for t in r["router_times"]])
    overall_s   = mean([t for r in results for t in r["specialist_times"]])
    overall_sum = mean([t for r in results for t in r["summary_times"]])
    overall_t   = mean(all_times)
    overall_c   = mean(all_composite) * 100
    print(
        f"{'':6} {'OVERALL':15} {'':50} | "
        f"{overall_r:>6.2f}s {overall_s:>7.2f}s {overall_sum:>6.2f}s {overall_t:>6.2f}s | "
        f"{'':>5} {'':>5} {'':>5} {overall_c:>5.0f}%"
    )
    print(f"{'='*W}\n")


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER  (imported lazily – saves ~200 MB peak RSS during inference)
# ──────────────────────────────────────────────────────────────────────────────

def _acc_fill(val: float):
    from openpyxl.styles import PatternFill
    if val >= 0.9:
        return PatternFill("solid", start_color="A6E3A1", end_color="A6E3A1")
    if val >= 0.5:
        return PatternFill("solid", start_color="FAB387", end_color="FAB387")
    return PatternFill("solid", start_color="F38BA8", end_color="F38BA8")


def _write_header_cell(ws, row: int, col: int, value: str) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=True, color="CDD6F4", size=11)
    cell.fill      = PatternFill("solid", start_color="1E1E2E", end_color="1E1E2E")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(
        left=Side(style="thin", color="45475A"), right=Side(style="thin", color="45475A"),
        top=Side(style="thin", color="45475A"),  bottom=Side(style="thin", color="45475A"),
    )


def _write_body_cell(ws, row: int, col: int, value: Any,
                     fill=None, fmt: Optional[str] = None,
                     align: str = "left") -> None:
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", color="CDD6F4", size=10)
    cell.fill      = fill or PatternFill()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = Border(
        left=Side(style="thin", color="45475A"), right=Side(style="thin", color="45475A"),
        top=Side(style="thin", color="45475A"),  bottom=Side(style="thin", color="45475A"),
    )
    if fmt:
        cell.number_format = fmt


def _setup_sheet(ws, headers: list[tuple[str, int]]) -> None:
    from openpyxl.utils import get_column_letter
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for ci, (hdr, width) in enumerate(headers, 1):
        _write_header_cell(ws, 1, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 30


def _write_raw_sheet(wb, all_raw: list[dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = "Raw Observations"

    headers = [
        ("Model", 22), ("TC ID", 7), ("Category", 14), ("Rep", 5),
        ("Question", 52), ("Expected Tool", 18), ("Detected Tool", 18),
        ("Router OK", 9), ("Router (s)", 10), ("Router Output", 45),
        ("Fn OK", 7), ("Spec (s)", 10), ("Specialist Output", 65),
        ("Param OK", 9), ("Summary (s)", 12), ("Summary Output", 45),
        ("Composite", 11), ("Total (s)", 10),
    ]
    _setup_sheet(ws, headers)

    from openpyxl.styles import PatternFill
    BEST = PatternFill("solid", start_color="A6E3A1", end_color="A6E3A1")
    FAIL = PatternFill("solid", start_color="F38BA8", end_color="F38BA8")

    bool_keys   = {"Router_Correct", "Fn_Correct", "Param_Correct"}
    time_keys   = {"Router_Time_s", "Specialist_Time_s", "Summary_Time_s", "Total_Time_s"}
    raw_keys    = [
        "Model", "TC_ID", "TC_Category", "Repetition",
        "Question", "Expected_Tool", "Detected_Tool",
        "Router_Correct", "Router_Time_s", "Router_Output",
        "Fn_Correct", "Specialist_Time_s", "Specialist_Output",
        "Param_Correct", "Summary_Time_s", "Summary_Output",
        "Composite_Acc", "Total_Time_s",
    ]

    for ri, row in enumerate(all_raw, 2):
        for ci, key in enumerate(raw_keys, 1):
            val = row.get(key, "")
            if key == "Composite_Acc":
                _write_body_cell(ws, ri, ci, val, fill=_acc_fill(float(val)),
                                 fmt="0.00", align="center")
            elif key in bool_keys:
                _write_body_cell(ws, ri, ci, "OK" if val else "FAIL",
                                 fill=BEST if val else FAIL, align="center")
            elif key in time_keys:
                _write_body_cell(ws, ri, ci, val, fmt="0.000", align="right")
            elif key == "Repetition":
                _write_body_cell(ws, ri, ci, val, align="center")
            else:
                _write_body_cell(ws, ri, ci, str(val) if val else "")


def _write_model_summary_sheet(wb, df) -> None:
    import numpy as np
    ws = wb.create_sheet("Model Summary")
    ws.sheet_view.showGridLines = False

    mg = df.groupby("Model").agg(
        avg_total_s    = ("Total_Time_s",      "mean"),
        median_total_s = ("Total_Time_s",      "median"),
        p95_total_s    = ("Total_Time_s",      lambda x: float(np.percentile(x, 95))),
        std_total_s    = ("Total_Time_s",      "std"),
        avg_router_s   = ("Router_Time_s",     "mean"),
        avg_spec_s     = ("Specialist_Time_s", "mean"),
        avg_sum_s      = ("Summary_Time_s",    "mean"),
        avg_composite  = ("Composite_Acc",     "mean"),
        avg_router_acc = ("Router_Correct",    "mean"),
        avg_fn_acc     = ("Fn_Correct",        "mean"),
        avg_param_acc  = ("Param_Correct",     "mean"),
        total_obs      = ("Repetition",        "count"),
    ).reset_index()

    mg["rank_lat"] = mg["avg_total_s"].rank()
    mg["rank_acc"] = mg["avg_composite"].rank(ascending=False)
    mg["rank"]     = (mg["rank_lat"] + mg["rank_acc"]).rank()
    mg = mg.sort_values("rank").reset_index(drop=True)

    headers = [
        ("Rank", 6), ("Model", 30), ("Avg Total(s)", 13), ("Median(s)", 10),
        ("P95(s)", 9), ("Std(s)", 9), ("Router(s)", 10), ("Spec(s)", 9),
        ("Summ(s)", 9), ("Composite", 11), ("Router Acc", 11),
        ("Fn Acc", 9), ("Param Acc", 11), ("Obs", 6),
    ]
    _setup_sheet(ws, headers)

    keys = [
        "rank", "Model", "avg_total_s", "median_total_s", "p95_total_s", "std_total_s",
        "avg_router_s", "avg_spec_s", "avg_sum_s", "avg_composite",
        "avg_router_acc", "avg_fn_acc", "avg_param_acc", "total_obs",
    ]
    fmts = [
        "0", "@", "0.000", "0.000", "0.000", "0.000", "0.000", "0.000", "0.000",
        "0.00%", "0.00%", "0.00%", "0.00%", "0",
    ]

    from openpyxl.styles import PatternFill
    BEST = PatternFill("solid", start_color="A6E3A1", end_color="A6E3A1")
    best_acc = mg["avg_composite"].max()
    best_lat = mg["avg_total_s"].min()

    for ri, (_, row) in enumerate(mg.iterrows(), 2):
        for ci, (key, fmt) in enumerate(zip(keys, fmts), 1):
            val  = row[key]
            fill = None
            al   = "left" if key == "Model" else "center"
            if key == "avg_composite":
                fill = BEST if val == best_acc else _acc_fill(val)
            elif key == "avg_total_s" and val == best_lat:
                fill = BEST
            _write_body_cell(ws, ri, ci, val, fill=fill, fmt=fmt, align=al)


def _write_per_tc_sheet(wb, df) -> None:
    ws = wb.create_sheet("Per-TC Breakdown")
    ws.sheet_view.showGridLines = False

    grp = df.groupby(["Model", "TC_ID", "TC_Category"]).agg(
        avg_total_s   = ("Total_Time_s",   "mean"),
        avg_composite = ("Composite_Acc",  "mean"),
        avg_router    = ("Router_Correct", "mean"),
        avg_fn        = ("Fn_Correct",     "mean"),
        avg_param     = ("Param_Correct",  "mean"),
        reps          = ("Repetition",     "count"),
    ).reset_index().sort_values(["Model", "TC_ID"])

    headers = [
        ("Model", 22), ("TC ID", 7), ("Category", 14), ("Avg Total(s)", 13),
        ("Composite", 12), ("Router", 9), ("Fn", 7), ("Param", 9), ("Reps", 6),
    ]
    _setup_sheet(ws, headers)

    keys = [
        "Model", "TC_ID", "TC_Category", "avg_total_s", "avg_composite",
        "avg_router", "avg_fn", "avg_param", "reps",
    ]
    fmts = ["@", "@", "@", "0.000", "0.00%", "0.00%", "0.00%", "0.00%", "0"]

    for ri, (_, row) in enumerate(grp.iterrows(), 2):
        for ci, (key, fmt) in enumerate(zip(keys, fmts), 1):
            val  = row[key]
            fill = _acc_fill(val) if key == "avg_composite" else None
            al   = "left" if key in ("Model", "TC_Category") else "center"
            _write_body_cell(ws, ri, ci, val, fill=fill, fmt=fmt, align=al)


def _write_per_category_sheet(wb, df) -> None:
    ws = wb.create_sheet("Per-Category")
    ws.sheet_view.showGridLines = False

    grp = df.groupby(["Model", "TC_Category"]).agg(
        avg_total_s   = ("Total_Time_s",   "mean"),
        avg_composite = ("Composite_Acc",  "mean"),
        avg_router    = ("Router_Correct", "mean"),
        avg_fn        = ("Fn_Correct",     "mean"),
        count         = ("Repetition",     "count"),
    ).reset_index().sort_values(["TC_Category", "Model"])

    headers = [
        ("Category", 14), ("Model", 26), ("Avg Total(s)", 13),
        ("Composite", 12), ("Router Acc", 11), ("Fn Acc", 9), ("Obs", 6),
    ]
    _setup_sheet(ws, headers)

    keys = ["TC_Category", "Model", "avg_total_s", "avg_composite", "avg_router", "avg_fn", "count"]
    fmts = ["@", "@", "0.000", "0.00%", "0.00%", "0.00%", "0"]

    for ri, (_, row) in enumerate(grp.iterrows(), 2):
        for ci, (key, fmt) in enumerate(zip(keys, fmts), 1):
            val  = row[key]
            fill = _acc_fill(val) if key == "avg_composite" else None
            al   = "left" if key in ("TC_Category", "Model") else "center"
            _write_body_cell(ws, ri, ci, val, fill=fill, fmt=fmt, align=al)


def _write_scatter_sheet(wb, df) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.chart.series import SeriesLabel
    import numpy as np

    ws = wb.create_sheet("Acc vs Latency")
    ws.sheet_view.showGridLines = False

    mg = df.groupby("Model").agg(
        avg_total_s   = ("Total_Time_s",  "mean"),
        avg_composite = ("Composite_Acc", "mean"),
    ).reset_index()

    for ci, h in enumerate(["Model", "Avg Total (s)", "Composite Acc (%)"], 1):
        _write_header_cell(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = 26

    for ri, (_, row) in enumerate(mg.iterrows(), 2):
        _write_body_cell(ws, ri, 1, row["Model"])
        _write_body_cell(ws, ri, 2, row["avg_total_s"],         fmt="0.000", align="center")
        _write_body_cell(ws, ri, 3, row["avg_composite"] * 100, fmt="0.00",  align="center")

    COLORS = [
        "818CF8", "C084FC", "34D399", "FB923C", "F472B6",
        "60A5FA", "FBBF24", "2DD4BF", "F87171", "4ADE80",
    ]
    scatter = ScatterChart()
    scatter.title        = "Accuracy vs Latency  (top-left = best)"
    scatter.style        = 10
    scatter.x_axis.title = "Avg Total Latency (s)"
    scatter.y_axis.title = "Composite Accuracy (%)"
    scatter.x_axis.numFmt = "0.0"
    scatter.y_axis.numFmt = "0"
    scatter.width  = 26
    scatter.height = 18

    for i, (_, row) in enumerate(mg.iterrows()):
        dr  = i + 2
        xv  = Reference(ws, min_col=2, min_row=dr, max_row=dr)
        yv  = Reference(ws, min_col=3, min_row=dr, max_row=dr)
        s   = Series(yv, xv)
        s.title = SeriesLabel(v=ws.cell(row=dr, column=1).value or f"Model {i+1}")
        col = COLORS[i % len(COLORS)]
        s.marker.symbol = "circle"
        s.marker.size   = 14
        s.marker.graphicalProperties.solidFill          = col
        s.marker.graphicalProperties.line.solidFill     = col
        s.graphicalProperties.line.noFill               = True
        scatter.series.append(s)

    ws.add_chart(scatter, "E2")


def build_excel(all_raw: list[dict[str, Any]], output_path: str) -> None:
    """Build the results Excel workbook from the flat list of raw observations."""
    import pandas as pd
    import openpyxl

    df = pd.DataFrame(all_raw)
    wb = openpyxl.Workbook()

    _write_raw_sheet(wb, all_raw)
    _write_model_summary_sheet(wb, df)
    _write_per_tc_sheet(wb, df)
    _write_per_category_sheet(wb, df)
    _write_scatter_sheet(wb, df)

    wb.save(output_path)
    print(f"\n  Excel saved -> {output_path}")


# ──────────────────────────────────────────────────────────────────────────────
# PUBLICATION PLOTS  (also imported lazily)
# ──────────────────────────────────────────────────────────────────────────────

# Colorblind-safe palette (Wong 2011)
_PALETTE = [
    "#0072B2", "#E69F00", "#009E73", "#CC79A7",
    "#56B4E9", "#D55E00", "#F0E442", "#000000",
]

_STAGE_COLORS = {
    "Router":     "#0072B2",
    "Specialist": "#E69F00",
    "Summarizer": "#009E73",
}

_METRIC_STYLES = [
    ("Router",  "avg_router_acc", "///",   "#0072B2"),
    ("Fn Call", "avg_fn_acc",     "\\\\\\","#E69F00"),
    ("Params",  "avg_param_acc",  "...",   "#009E73"),
]


def _short(model_id: str) -> str:
    return model_id.split("/")[-1]


def _pub_style() -> None:
    import matplotlib.pyplot as plt
    plt.rcParams.update({
        "font.family":       "sans-serif",
        "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
        "font.size":         9,
        "axes.titlesize":    10,
        "axes.labelsize":    9,
        "xtick.labelsize":   8,
        "ytick.labelsize":   8,
        "legend.fontsize":   8,
        "axes.spines.top":   False,
        "axes.spines.right": False,
        "axes.linewidth":    0.8,
        "axes.grid":         True,
        "grid.linestyle":    "--",
        "grid.linewidth":    0.4,
        "grid.alpha":        0.5,
        "grid.color":        "#cccccc",
        "xtick.direction":   "out",
        "ytick.direction":   "out",
        "xtick.major.size":  3,
        "ytick.major.size":  3,
        "xtick.major.width": 0.8,
        "ytick.major.width": 0.8,
        "figure.dpi":        150,
        "savefig.dpi":       300,
        "savefig.bbox":      "tight",
        "savefig.pad_inches": 0.05,
        "lines.linewidth":   1.5,
        "lines.markersize":  6,
        "legend.frameon":    True,
        "legend.framealpha": 0.9,
        "legend.edgecolor":  "#cccccc",
        "legend.handlelength": 1.5,
    })


def _savefig(fig, out_dir: str, name: str, fmt: str) -> None:
    import matplotlib.pyplot as plt
    path = os.path.join(out_dir, f"{name}.{fmt}")
    fig.savefig(path)
    plt.close(fig)
    print(f"  [plot] {path}")


def _fig_acc_vs_latency(mg, color_map: dict, out_dir: str, fmt: str) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MultipleLocator

    fig, ax = plt.subplots(figsize=(4.5, 3.5))

    for model, row in mg.iterrows():
        col = color_map[model]
        ax.errorbar(
            row["avg_total_s"], row["avg_composite"] * 100,
            xerr=row["std_total_s"],
            fmt="o", color=col, markersize=7,
            capsize=3, capthick=0.8, elinewidth=0.8,
            label=_short(model), zorder=3,
        )
        ax.annotate(
            _short(model),
            xy=(row["avg_total_s"], row["avg_composite"] * 100),
            xytext=(5, 3), textcoords="offset points",
            fontsize=7, color=col,
        )

    ax.axhline(mg["avg_composite"].max() * 100, color="#aaaaaa", ls=":", lw=0.8, zorder=1)
    ax.axvline(mg["avg_total_s"].min(),          color="#aaaaaa", ls=":", lw=0.8, zorder=1)
    ax.set_xlabel("Average Total Latency (s)")
    ax.set_ylabel("Composite Accuracy (%)")
    ax.set_title("Accuracy vs. Latency\n(top-left = best)")
    ax.yaxis.set_major_locator(MultipleLocator(10))
    ax.set_ylim(0, 105)
    ax.set_xlim(left=0)
    fig.tight_layout()
    _savefig(fig, out_dir, "acc_vs_latency", fmt)


def _fig_latency_breakdown(mg, labels: list[str], out_dir: str, fmt: str) -> None:
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(max(3.5, len(labels) * 0.9 + 1.5), 3.5))
    x      = np.arange(len(labels))
    bottom = np.zeros(len(labels))

    for stage_label, col_key, color in [
        ("Router",     "avg_router_s", _STAGE_COLORS["Router"]),
        ("Specialist", "avg_spec_s",   _STAGE_COLORS["Specialist"]),
        ("Summarizer", "avg_sum_s",    _STAGE_COLORS["Summarizer"]),
    ]:
        vals = mg[col_key].values
        ax.bar(x, vals, 0.55, bottom=bottom, label=stage_label,
               color=color, edgecolor="white", linewidth=0.5, zorder=3)
        bottom += vals

    ax.errorbar(x, mg["avg_total_s"].values, yerr=mg["std_total_s"].values,
                fmt="none", color="#333333", capsize=3, capthick=0.8, elinewidth=0.8, zorder=4)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=20, ha="right")
    ax.set_ylabel("Latency (s)")
    ax.set_title("Pipeline Latency Breakdown\nper Stage")
    ax.legend(loc="upper right")
    ax.set_ylim(bottom=0)
    fig.tight_layout()
    _savefig(fig, out_dir, "latency_breakdown", fmt)


def _fig_accuracy_breakdown(mg, labels: list[str], out_dir: str, fmt: str) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MultipleLocator

    fig, ax = plt.subplots(figsize=(max(3.5, len(labels) * 1.2 + 1.5), 3.5))
    x      = np.arange(len(labels))
    w      = 0.22
    offs   = np.array([-w, 0, w])

    for (lbl, key, hatch, color), offset in zip(_METRIC_STYLES, offs):
        vals = mg[key].values * 100
        bars = ax.bar(x + offset, vals, w, label=lbl, color=color, alpha=0.85,
                      hatch=hatch, edgecolor="white", linewidth=0.4, zorder=3)
        for bar, v in zip(bars, vals):
            if v > 5:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1,
                        f"{v:.0f}", ha="center", va="bottom", fontsize=6.5)

    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=20, ha="right")
    ax.set_ylabel("Accuracy (%)")
    ax.set_title("Accuracy by Component")
    ax.set_ylim(0, 115)
    ax.legend(loc="upper right")
    ax.yaxis.set_major_locator(MultipleLocator(20))
    ax.axhline(100, color="#cccccc", ls="--", lw=0.6, zorder=1)
    fig.tight_layout()
    _savefig(fig, out_dir, "accuracy_breakdown", fmt)


def _fig_category_heatmap(cat_pivot, labels: list[str], out_dir: str, fmt: str):
    import matplotlib.pyplot as plt

    categories = cat_pivot.columns.tolist()
    fig, ax = plt.subplots(figsize=(max(4, len(categories) * 0.9 + 1.5),
                                    max(2.5, len(labels) * 0.6 + 1.0)))
    im = ax.imshow(cat_pivot.values, aspect="auto", cmap="RdYlGn", vmin=0, vmax=100)

    ax.set_xticks(np.arange(len(categories)))
    ax.set_yticks(np.arange(len(labels)))
    ax.set_xticklabels(categories, rotation=35, ha="right", fontsize=8)
    ax.set_yticklabels(labels, fontsize=8)
    ax.tick_params(length=0)
    ax.spines[:].set_visible(False)

    for i in range(len(labels)):
        for j in range(len(categories)):
            v = cat_pivot.values[i, j]
            if not np.isnan(v):
                tc = "white" if v < 40 or v > 80 else "black"
                ax.text(j, i, f"{v:.0f}", ha="center", va="center",
                        fontsize=7.5, color=tc, fontweight="bold")

    cbar = fig.colorbar(im, ax=ax, fraction=0.03, pad=0.02)
    cbar.set_label("Composite Accuracy (%)", fontsize=8)
    cbar.ax.tick_params(labelsize=7)
    ax.set_title("Composite Accuracy by Model × Task Category")
    fig.tight_layout()
    _savefig(fig, out_dir, "category_heatmap", fmt)


def _fig_per_tc_strip(tc_avg, model_col: list[str], labels: list[str],
                      color_map: dict, rng, out_dir: str, fmt: str):
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MultipleLocator

    bp_data = [tc_avg[tc_avg["Model"] == m]["Composite_Acc"].values * 100 for m in model_col]
    jitter_w = 0.12

    fig, ax = plt.subplots(figsize=(max(4, len(labels) * 1.1 + 1), 3.5))
    bp = ax.boxplot(bp_data, positions=np.arange(len(model_col)), widths=0.3,
                    patch_artist=True, showfliers=False,
                    medianprops=dict(color="black", linewidth=1.5),
                    whiskerprops=dict(linewidth=0.8),
                    capprops=dict(linewidth=0.8),
                    boxprops=dict(linewidth=0.8))

    for patch, model in zip(bp["boxes"], model_col):
        patch.set_facecolor(color_map[model])
        patch.set_alpha(0.25)

    for pos, model in zip(np.arange(len(model_col)), model_col):
        vals   = tc_avg[tc_avg["Model"] == model]["Composite_Acc"].values * 100
        jitter = rng.uniform(-jitter_w, jitter_w, len(vals))
        ax.scatter(pos + jitter, vals, s=18, color=color_map[model],
                   alpha=0.75, zorder=4, linewidths=0)

    ax.set_xticks(np.arange(len(labels)))
    ax.set_xticklabels(labels, rotation=20, ha="right")
    ax.set_ylabel("Composite Accuracy (%) — per TC")
    ax.set_title("Per-Question Accuracy Distribution")
    ax.set_ylim(-5, 108)
    ax.axhline(100, color="#cccccc", ls="--", lw=0.6)
    ax.yaxis.set_major_locator(MultipleLocator(20))
    fig.tight_layout()
    _savefig(fig, out_dir, "per_tc_strip", fmt)
    return bp_data


def _fig_latency_box(df, model_col: list[str], labels: list[str],
                     color_map: dict, out_dir: str, fmt: str) -> None:
    import matplotlib.pyplot as plt

    lat_data = [df[df["Model"] == m]["Total_Time_s"].values for m in model_col]
    fig, ax  = plt.subplots(figsize=(max(4, len(labels) * 1.1 + 1), 3.5))
    bp2 = ax.boxplot(lat_data, positions=np.arange(len(model_col)), widths=0.4,
                     patch_artist=True, showfliers=True,
                     flierprops=dict(marker=".", markersize=3, alpha=0.4),
                     medianprops=dict(color="black", linewidth=1.5),
                     whiskerprops=dict(linewidth=0.8),
                     capprops=dict(linewidth=0.8),
                     boxprops=dict(linewidth=0.8))

    for patch, model in zip(bp2["boxes"], model_col):
        patch.set_facecolor(color_map[model])
        patch.set_alpha(0.55)

    ax.set_xticks(np.arange(len(model_col)))
    ax.set_xticklabels(labels, rotation=20, ha="right")
    ax.set_ylabel("Total Latency (s)")
    ax.set_title("Total Latency Distribution\n(all observations)")
    ax.set_ylim(bottom=0)
    fig.tight_layout()
    _savefig(fig, out_dir, "latency_box", fmt)


def _fig_summary_panel(mg, cat_pivot, tc_avg, model_col: list[str],
                       labels: list[str], color_map: dict,
                       bp_data: list, rng, out_dir: str, fmt: str) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.gridspec import GridSpec
    from matplotlib.ticker import MultipleLocator

    jitter_w = 0.12
    fig = plt.figure(figsize=(10, 6.5))
    gs  = GridSpec(2, 3, figure=fig, hspace=0.52, wspace=0.38)

    ax_scatter = fig.add_subplot(gs[0, 0])
    ax_lat_brk = fig.add_subplot(gs[0, 1])
    ax_acc_brk = fig.add_subplot(gs[0, 2])
    ax_heatmap = fig.add_subplot(gs[1, 0:2])
    ax_strip   = fig.add_subplot(gs[1, 2])
    categories = cat_pivot.columns.tolist()

    for model, row in mg.iterrows():
        col = color_map[model]
        ax_scatter.errorbar(
            row["avg_total_s"], row["avg_composite"] * 100,
            xerr=row["std_total_s"],
            fmt="o", color=col, markersize=6,
            capsize=2.5, capthick=0.7, elinewidth=0.7,
            label=_short(model), zorder=3,
        )
        ax_scatter.annotate(_short(model),
                            xy=(row["avg_total_s"], row["avg_composite"] * 100),
                            xytext=(4, 2), textcoords="offset points",
                            fontsize=6, color=col)
    ax_scatter.set_xlabel("Avg Latency (s)")
    ax_scatter.set_ylabel("Composite Acc. (%)")
    ax_scatter.set_ylim(0, 108)
    ax_scatter.set_xlim(left=0)
    ax_scatter.yaxis.set_major_locator(MultipleLocator(20))

    x_b = np.arange(len(labels))
    bot = np.zeros(len(labels))
    for stage_label, col_key, color in [
        ("Router",     "avg_router_s", _STAGE_COLORS["Router"]),
        ("Specialist", "avg_spec_s",   _STAGE_COLORS["Specialist"]),
        ("Summarizer", "avg_sum_s",    _STAGE_COLORS["Summarizer"]),
    ]:
        vals = mg[col_key].values
        ax_lat_brk.bar(x_b, vals, 0.5, bottom=bot, label=stage_label,
                       color=color, edgecolor="white", linewidth=0.4, zorder=3)
        bot += vals
    ax_lat_brk.set_xticks(x_b)
    ax_lat_brk.set_xticklabels(labels, rotation=22, ha="right", fontsize=7)
    ax_lat_brk.set_ylabel("Latency (s)")
    ax_lat_brk.legend(fontsize=6, loc="upper right")
    ax_lat_brk.set_ylim(bottom=0)

    x_c  = np.arange(len(labels))
    w_c  = 0.22
    offs = np.array([-w_c, 0, w_c])
    for (lbl, key, hatch, color), off in zip(_METRIC_STYLES, offs):
        ax_acc_brk.bar(x_c + off, mg[key].values * 100, w_c, label=lbl,
                       color=color, alpha=0.85, hatch=hatch,
                       edgecolor="white", linewidth=0.3, zorder=3)
    ax_acc_brk.set_xticks(x_c)
    ax_acc_brk.set_xticklabels(labels, rotation=22, ha="right", fontsize=7)
    ax_acc_brk.set_ylabel("Accuracy (%)")
    ax_acc_brk.set_ylim(0, 118)
    ax_acc_brk.legend(fontsize=6, loc="upper right")
    ax_acc_brk.axhline(100, color="#cccccc", ls="--", lw=0.5)
    ax_acc_brk.yaxis.set_major_locator(MultipleLocator(25))

    im2 = ax_heatmap.imshow(cat_pivot.values, aspect="auto",
                             cmap="RdYlGn", vmin=0, vmax=100)
    ax_heatmap.set_xticks(np.arange(len(categories)))
    ax_heatmap.set_yticks(np.arange(len(labels)))
    ax_heatmap.set_xticklabels(categories, rotation=28, ha="right", fontsize=7)
    ax_heatmap.set_yticklabels(labels, fontsize=7)
    ax_heatmap.tick_params(length=0)
    ax_heatmap.spines[:].set_visible(False)
    for i in range(len(labels)):
        for j in range(len(categories)):
            v = cat_pivot.values[i, j]
            if not np.isnan(v):
                tc_c = "white" if v < 40 or v > 80 else "black"
                ax_heatmap.text(j, i, f"{v:.0f}", ha="center", va="center",
                                fontsize=6.5, color=tc_c, fontweight="bold")
    cbar2 = fig.colorbar(im2, ax=ax_heatmap, fraction=0.025, pad=0.02)
    cbar2.set_label("Acc. (%)", fontsize=7)
    cbar2.ax.tick_params(labelsize=6)

    bp3 = ax_strip.boxplot(bp_data, positions=np.arange(len(model_col)), widths=0.3,
                           patch_artist=True, showfliers=False,
                           medianprops=dict(color="black", linewidth=1.2),
                           whiskerprops=dict(linewidth=0.7),
                           capprops=dict(linewidth=0.7),
                           boxprops=dict(linewidth=0.7))
    for patch, model in zip(bp3["boxes"], model_col):
        patch.set_facecolor(color_map[model])
        patch.set_alpha(0.25)
    for pos, model in zip(np.arange(len(model_col)), model_col):
        vals   = tc_avg[tc_avg["Model"] == model]["Composite_Acc"].values * 100
        jitter = rng.uniform(-jitter_w, jitter_w, len(vals))
        ax_strip.scatter(pos + jitter, vals, s=14, color=color_map[model],
                         alpha=0.75, zorder=4, linewidths=0)
    ax_strip.set_xticks(np.arange(len(labels)))
    ax_strip.set_xticklabels(labels, rotation=22, ha="right", fontsize=7)
    ax_strip.set_ylabel("Acc. (%) per TC")
    ax_strip.set_ylim(-5, 108)
    ax_strip.axhline(100, color="#cccccc", ls="--", lw=0.5)
    ax_strip.yaxis.set_major_locator(MultipleLocator(25))

    for ax_p, plabel in zip([ax_scatter, ax_lat_brk, ax_acc_brk, ax_heatmap, ax_strip],
                             ["(a)", "(b)", "(c)", "(d)", "(e)"]):
        ax_p.set_title(plabel, loc="left", fontweight="bold", fontsize=9, pad=2)

    fig.suptitle("MLX Model Benchmark — Pipeline Accuracy & Latency",
                 fontsize=11, fontweight="bold", y=1.01)
    fig.tight_layout()
    _savefig(fig, out_dir, "summary_panel", fmt)


def plot_results(all_raw: list[dict[str, Any]], model_ids: list[str],
                 out_dir: str = ".", fmt: str = "pdf") -> None:
    import pandas as pd
    import matplotlib
    matplotlib.use("Agg")

    os.makedirs(out_dir, exist_ok=True)
    _pub_style()

    df        = pd.DataFrame(all_raw)
    model_col = df["Model"].unique().tolist()
    color_map = {m: _PALETTE[i % len(_PALETTE)] for i, m in enumerate(model_col)}
    labels    = [_short(m) for m in model_col]

    mg = df.groupby("Model").agg(
        avg_total_s    = ("Total_Time_s",      "mean"),
        std_total_s    = ("Total_Time_s",      "std"),
        avg_router_s   = ("Router_Time_s",     "mean"),
        avg_spec_s     = ("Specialist_Time_s", "mean"),
        avg_sum_s      = ("Summary_Time_s",    "mean"),
        avg_composite  = ("Composite_Acc",     "mean"),
        avg_router_acc = ("Router_Correct",    "mean"),
        avg_fn_acc     = ("Fn_Correct",        "mean"),
        avg_param_acc  = ("Param_Correct",     "mean"),
    ).reindex(model_col)

    cat_pivot = (
        df.groupby(["Model", "TC_Category"])["Composite_Acc"]
        .mean()
        .unstack("TC_Category")
        .reindex(model_col)
        * 100
    )

    tc_avg = (
        df.groupby(["Model", "TC_ID"])["Composite_Acc"]
        .mean()
        .reset_index()
    )

    rng = np.random.default_rng(42)

    _fig_acc_vs_latency(mg, color_map, out_dir, fmt)
    _fig_latency_breakdown(mg, labels, out_dir, fmt)
    _fig_accuracy_breakdown(mg, labels, out_dir, fmt)
    _fig_category_heatmap(cat_pivot, labels, out_dir, fmt)
    bp_data = _fig_per_tc_strip(tc_avg, model_col, labels, color_map, rng, out_dir, fmt)
    _fig_latency_box(df, model_col, labels, color_map, out_dir, fmt)
    _fig_summary_panel(mg, cat_pivot, tc_avg, model_col, labels,
                       color_map, bp_data, rng, out_dir, fmt)

    print(f"\n  [plots] All figures written to: {out_dir}/")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    global REPETITIONS_PER_Q

    parser = argparse.ArgumentParser(description="MLX multi-model pipeline benchmark")
    parser.add_argument("mode",     nargs="?",  default="",
                        help="'quick' for a 1-rep smoke-test")
    parser.add_argument("--reps",   type=int,   default=None,
                        help=f"Repetitions per question (default: {REPETITIONS_PER_Q})")
    parser.add_argument("--output", default=None,
                        help="Output Excel path (auto-timestamped if omitted)")
    parser.add_argument("--plots",  default=None, metavar="DIR",
                        help="Write publication plots into DIR (e.g. --plots figures/)")
    parser.add_argument("--fmt",    default="pdf", choices=["pdf", "svg", "png"],
                        help="Plot file format (default: pdf)")
    args = parser.parse_args()

    if args.mode == "quick":
        REPETITIONS_PER_Q = 1
    elif args.reps:
        REPETITIONS_PER_Q = args.reps

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or f"benchmark_mlx_{timestamp}.xlsx"
    ckpt_path   = output_path.replace(".xlsx", "_checkpoint.csv")

    total_calls = len(MLX_MODELS) * len(TEST_CASES) * REPETITIONS_PER_Q

    print(f"\nMLX Benchmark  (M1 Pro / 8 GB optimised)")
    print(f"  Backend root : {_BACKEND_ROOT}")
    print(f"  Models       : {len(MLX_MODELS)}")
    print(f"  Test cases   : {len(TEST_CASES)}")
    print(f"  Reps/Q       : {REPETITIONS_PER_Q}")
    print(f"  Total calls  : {total_calls}")
    print(f"  Output Excel : {output_path}")
    print(f"  Checkpoint   : {ckpt_path}  (flushed after every rep)")

    if not ROUTER_PROMPT:
        print("\n  WARNING: router_prompt.txt empty/missing – router stage has blank system prompt.")
    if not SUMMARY_PROMPT:
        print("  WARNING: summary_prompt.txt empty/missing – summarizer stage will be skipped.\n")

    ckpt_writer, ckpt_fh = open_checkpoint(ckpt_path)

    try:
        # FIX 3: removed the stale single-model branch with the misleading comment.
        # run_benchmark always writes to checkpoint; main() always reads back from
        # CSV after the loop, so there is one consistent code path for all cases.
        for model_id in MLX_MODELS:
            per_q = run_benchmark(model_id, REPETITIONS_PER_Q, ckpt_writer, ckpt_fh)
            print_results(model_id, per_q)

            if len(MLX_MODELS) > 1:
                print(f"\n  [memory] Freeing model '{model_id.split('/')[-1]}' before next load …")
                free_model_memory()

    finally:
        ckpt_fh.close()

    # Re-read checkpoint as single source of truth for Excel / plots.
    # This avoids keeping a second copy of all_raw in RAM during inference.
    import csv as _csv
    all_raw: list[dict[str, Any]] = []
    with open(ckpt_path, newline="", encoding="utf-8") as f:
        reader = _csv.DictReader(f)
        all_raw = list(reader)

    # Cast numeric fields back from strings
    _numeric_fields = {
        "Router_Correct", "Fn_Correct", "Param_Correct",
        "Router_Time_s", "Specialist_Time_s", "Summary_Time_s",
        "Composite_Acc", "Total_Time_s", "Repetition",
    }
    for row in all_raw:
        for field in _numeric_fields:
            if field in row and row[field] != "":
                try:
                    row[field] = float(row[field])
                except ValueError:
                    pass

    build_excel(all_raw, output_path)

    if args.plots is not None:
        print(f"\n  Generating publication plots -> {args.plots}/")
        plot_results(all_raw, MLX_MODELS, out_dir=args.plots, fmt=args.fmt)

    import pandas as pd
    df = pd.DataFrame(all_raw)
    lb = (
        df.groupby("Model")
        .agg(avg_latency_s=("Total_Time_s", "mean"),
             composite_acc=("Composite_Acc", "mean"))
        .round(3)
        .sort_values("composite_acc", ascending=False)
    )
    print("\n  Leaderboard")
    print("  " + "-" * 46)
    print(lb.to_string())
    print(f"\n  Done  Excel  -> {output_path}")
    print(f"        Checkpoint -> {ckpt_path}")


if __name__ == "__main__":
    main()