# backend/expense_benchmark.py
"""
expense_benchmark.py

Benchmarks two routing setups for the expense assistant:

1) single      -> one model does routing + specialization in one shot
2) dual        -> router model selects tool, specialist model emits function call
3) validated   -> post-process specialist output through validate_and_fix_params

Writes a crash-safe CSV checkpoint after every repetition and an Excel summary at the end.
Designed to preserve the same core reporting style as the current benchmark: latency, peak RAM/CPU,
and the paper-style metrics FSP / ACS / AVC / TaskAcc / CorrectRatio.

Typical usage:
    python expense_benchmark.py --mode both
    python expense_benchmark.py --mode single --reps 3
    python expense_benchmark.py --mode dual --models llama3-8b,phi3-mini
    python expense_benchmark.py --mode both --resume   # append to existing checkpoint
"""

from __future__ import annotations

import argparse
import ast
import csv
import gc
import json
import logging
import os
import re
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, Iterable, Optional, Tuple

import numpy as np
import pandas as pd
import psutil

# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------
TEMPERATURE = 0.1

# -----------------------------------------------------------------------------
# Logging
# -----------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger("expense_benchmark")

# -----------------------------------------------------------------------------
# Path resolution
# -----------------------------------------------------------------------------
def _resolve_backend_root() -> str:
    candidates = [
        os.getcwd(),
        os.path.join(os.getcwd(), "javascript_app", "backend"),
        os.path.join(os.getcwd(), "backend"),
        os.path.dirname(os.path.abspath(__file__)),
    ]
    for root in candidates:
        if os.path.isdir(os.path.join(root, "utils")):
            return root
    return os.path.dirname(os.path.abspath(__file__))


BACKEND_ROOT = _resolve_backend_root()
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

EXPERIMENTS_SHARED = os.path.join(BACKEND_ROOT, "experiments", "shared")
EXPERIMENTS_CONFIGS = os.path.join(BACKEND_ROOT, "experiments", "configs")
for p in [EXPERIMENTS_SHARED, EXPERIMENTS_CONFIGS]:
    if os.path.isdir(p) and p not in sys.path:
        sys.path.insert(0, p)

# -----------------------------------------------------------------------------
# Imports from project
# -----------------------------------------------------------------------------
from experiments.inference import generate, get_last_usage
from experiments.models import get_llamacpp_models
from experiments.memory import free_model_memory
from utils.llm_input_validation import validate_and_fix_params
from utils.tool_prompts import get_tool_prompt

# -----------------------------------------------------------------------------
# Prompt loading
# -----------------------------------------------------------------------------
def load_text_file(*relative_paths: str) -> str:
    for rel in relative_paths:
        path = os.path.join(BACKEND_ROOT, rel)
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
    return ""


ROUTER_PROMPT = load_text_file("utils/prompts/router_prompt.txt")
if not ROUTER_PROMPT:
    logger.info("Router prompt not found, using default prompt.")
    ROUTER_PROMPT = """You are an expert intent classifier for an expense analysis assistant.
Your job is to determine which ONE tool is best suited to answer the user's question.

## Available Tools

1. **`plot_time_series`**
   - Use when user asks about: trends, spending over time, "past X months", "since 2023", date ranges.
   - Keywords: trend, over time, months, years, since.

2. **`plot_distribution`**
   - Use when user asks for: breakdown, distribution, proportions, pie chart, "how is X split", "excluding rent".
   - Keywords: breakdown, distribution, pie chart, split, exclude rent, without rent.

3. **`plot_comparison_bars`**
   - Use ONLY when comparing TWO different time periods (e.g. "Dec 2024 vs Dec 2025").
   - REQUIRES two distinct periods. If the user only mentions one month or one year, DO NOT use this tool.
   - Keywords: compare, vs, versus, difference between.

4. **`calculate_total`**
   - Use when user asks for: simple totals, sums, specific amounts "how much did I spend".
   - Keywords: how much, total, sum, cost.

5. **`get_top_expenses`**
   - Use when user asks for: biggest/largest expenses, top X items, most expensive.
   - Keywords: biggest, largest, top, most expensive, highest, excluding rent.

## Output Format
Output ONLY the tool name. Do not output anything else.
If the intent is unclear, default to `calculate_total`.

## Examples
User: "How much did I spend on food?"
Output: calculate_total

User: "Show me my spending trend for groceries"
Output: plot_time_series

User: "Compare food spending in 2024 and 2025"
Output: plot_comparison_bars

User: "What were my biggest expenses?"
Output: get_top_expenses

User: "Pie chart of transportation"
Output: plot_distribution

User: "Top 10 items without rent"
Output: get_top_expenses
"""


def build_single_agent_prompt(metadata: str, current_date: str) -> str:
    return f"""
You are an intelligent API function caller for an expense analysis system.

Your task has TWO steps:

STEP 1: Decide the BEST tool to answer the user's query.
STEP 2: Call that function with the correct parameters.

You MUST internally decide the tool first, but you MUST NOT output the tool name separately.

FINAL OUTPUT RULE:
- Output EXACTLY ONE line: a valid function call.
- DO NOT explain anything.
- DO NOT output JSON.
- DO NOT include backticks.
- DO NOT include any text before or after the function call.
- Use EXACT category names from the metadata below. If not an exact match, map it to the closest one.
- For both broad/major categories (e.g. "Food") and specific sub-categories (e.g. "grocery"), use category=.

--------------------------------------------------

## AVAILABLE TOOLS

1. plot_time_series(df, category=None, year=None, month=None, start_year=None, start_month=None, end_year=None, end_month=None, months=None)

2. plot_distribution(df, category=None, remarks=None, year=None, month=None, start_year=None, start_month=None, end_year=None, end_month=None, months=None, ignore_rent=False)

3. plot_comparison_bars(df, category=None, y1=None, m1=None, d1=None, y2=None, m2=None, d2=None)

4. calculate_total(df, category=None, remarks=None, year=None, month=None, day=None, start_year=None, start_month=None, end_year=None, end_month=None, months=None)

5. get_top_expenses(df, n=10, category=None, year=None, month=None, day=None, start_year=None, start_month=None, end_year=None, end_month=None, months=None, min_amount=None, ignore_rent=False)

--------------------------------------------------

## CONTEXT
{metadata}

Today: {current_date}

FINAL REMINDER:
Output ONLY the function call.
"""


# -----------------------------------------------------------------------------
# Categories metadata
# -----------------------------------------------------------------------------
def load_categories() -> dict:
    candidates = [
        os.path.join(BACKEND_ROOT, "..", "src", "utils", "categories.json"),
        os.path.join(BACKEND_ROOT, "src", "utils", "categories.json"),
        os.path.join(BACKEND_ROOT, "utils", "categories.json"),
    ]
    for path in candidates:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    return {"CATEGORY_COLORS": {}, "CATEGORY_MAPPING": {}}


CATEGORIES = load_categories()
def _deduplicate_category_mapping(mapping: dict) -> dict:
    seen: dict[str, str] = {}
    deduped: dict[str, str] = {}
    for original_key, value in mapping.items():
        lower_key = original_key.lower()
        if lower_key not in seen:
            seen[lower_key] = original_key
            deduped[original_key] = value
    return deduped

CATEGORY_MAPPING = _deduplicate_category_mapping(CATEGORIES.get("CATEGORY_MAPPING", {}))
MAJOR_CATEGORIES = list(CATEGORIES.get("CATEGORY_COLORS", {}).keys())

def build_metadata_text() -> str:
    all_categories = sorted(set(MAJOR_CATEGORIES) | set(CATEGORY_MAPPING.keys()))
    lines = ["### CATEGORIES (available options for 'category' argument):"]
    for cat in all_categories:
        lines.append(f"- {cat}")
    return "\n".join(lines)

METADATA_TEXT = build_metadata_text()

def build_validation_df() -> pd.DataFrame:
    rows = []
    for major in MAJOR_CATEGORIES:
        rows.append({"category": major, "major category": major})
    for cat, major in CATEGORY_MAPPING.items():
        rows.append({"category": cat, "major category": major})
    return pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)


VALIDATION_DF = build_validation_df()

# Import TEST_CASES from backend/experiments/test_cases.py explicitly (not experiments/configs/test_cases.py)
import importlib.util as _ilu
_tc_spec = _ilu.spec_from_file_location("backend_test_cases", os.path.join(BACKEND_ROOT, "experiments", "test_cases.py"))
_tc_mod = _ilu.module_from_spec(_tc_spec)
_tc_spec.loader.exec_module(_tc_mod)
TEST_CASES = _tc_mod.TEST_CASES


# -----------------------------------------------------------------------------
# Monitoring
# -----------------------------------------------------------------------------
class ResourceMonitor:
    def __init__(self, interval: float = 0.05):
        self.interval = interval
        self._stop_event = threading.Event()
        self.peak_cpu_percent = 0.0
        self.peak_ram_mb = 0.0
        self.process = psutil.Process(os.getpid())
        self.thread: Optional[threading.Thread] = None

    def _monitor(self) -> None:
        self.process.cpu_percent(interval=None)
        while not self._stop_event.is_set():
            try:
                cpu = self.process.cpu_percent(interval=None) / max(psutil.cpu_count() or 1, 1)
                ram_mb = self.process.memory_info().rss / (1024 * 1024)
                self.peak_cpu_percent = max(self.peak_cpu_percent, cpu)
                self.peak_ram_mb = max(self.peak_ram_mb, ram_mb)
            except Exception:
                pass
            self._stop_event.wait(self.interval)

    def __enter__(self):
        self._stop_event.clear()
        self.peak_cpu_percent = 0.0
        self.peak_ram_mb = 0.0
        self.thread = threading.Thread(target=self._monitor, daemon=True)
        self.thread.start()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._stop_event.set()
        if self.thread:
            self.thread.join()


# -----------------------------------------------------------------------------
# Parsing helpers
# -----------------------------------------------------------------------------
ALLOWED_TOOLS = {
    "plot_time_series",
    "plot_distribution",
    "plot_comparison_bars",
    "calculate_total",
    "get_top_expenses",
}


def strip_code_fences(text: str) -> str:
    if not text:
        return ""
    s = text.strip()
    if "```" in s:
        blocks = re.findall(r"```(?:python)?\s*(.*?)\s*```", s, re.DOTALL | re.IGNORECASE)
        if blocks:
            s = blocks[0].strip()
        else:
            s = s.replace("```python", "").replace("```", "").strip()
    return s.strip().strip("`")


def extract_first_tool_call(text: str) -> tuple[str, dict[str, Any], str]:
    """
    Returns (tool_name, kwargs, cleaned_text).
    Accepts raw function call, or wrapped text.
    """
    cleaned = strip_code_fences(text)
    if not cleaned:
        return "NONE", {}, cleaned

    call_match = re.search(
        r"(plot_time_series|plot_distribution|plot_comparison_bars|calculate_total|get_top_expenses)\s*\(",
        cleaned,
    )
    if not call_match:
        token = cleaned.split()[0].strip().strip("`'\"")
        return token if token in ALLOWED_TOOLS else token, {}, cleaned

    start = call_match.start(1)
    candidate = cleaned[start:].strip()

    try:
        tree = ast.parse(f"__x__ = {candidate}")
        call_node = None
        for node in ast.walk(tree):
            if isinstance(node, ast.Call):
                call_node = node
                break
        if call_node is None:
            return "NONE", {}, cleaned
        func = call_node.func
        tool_name = func.id if isinstance(func, ast.Name) else getattr(func, "attr", "NONE")

        kwargs: dict[str, Any] = {}
        for kw in call_node.keywords:
            if kw.arg is None:
                continue
            try:
                kwargs[kw.arg] = ast.literal_eval(kw.value)
            except Exception:
                try:
                    kwargs[kw.arg] = ast.unparse(kw.value)
                except Exception:
                    kwargs[kw.arg] = None
        return tool_name, kwargs, cleaned
    except Exception:
        tool_name = call_match.group(1)
        kwargs = {}
        return tool_name, kwargs, cleaned


def canonicalize_params(params: dict[str, Any]) -> dict[str, Any]:
    """Map major_category -> category for grading consistency."""
    out = dict(params or {})
    # Map major_category to category if present (for old expected_params in test cases)
    if "major_category" in out and "category" not in out:
        out["category"] = out.pop("major_category")
    if "major_category" in out and "category" in out:
        if out["category"] in (None, ""):
            out["category"] = out["major_category"]
        out.pop("major_category", None)
    return out


def params_to_call_str(params: dict[str, Any]) -> str:
    parts = []
    for k in sorted(params.keys()):
        v = params[k]
        if v is None:
            continue
        parts.append(f"{k}={repr(v)}")
    return ", ".join(parts)


def expected_set(tool: str, expected_params: dict[str, Any]) -> set[str]:
    s = {tool.lower()}
    for k, v in expected_params.items():
        s.add(f"{k}={str(v).lower()}")
    return s


def predicted_set(tool: str, params: dict[str, Any]) -> set[str]:
    s = {tool.lower()}
    for k, v in params.items():
        # FIX: filter the positional 'df' argument that AST parsing may capture.
        # It is never part of the expected parameter set and would inflate false negatives.
        if v is not None and k.lower() != "df":
            s.add(f"{k}={str(v).lower()}")
    return s


def _lower_dict(d: dict[str, Any]) -> dict[str, Any]:
    """Return a copy of *d* whose keys are lowercased.

    When two original keys collide after lowering, the first occurrence wins.
    This ensures that `dict.get(lowered_key)` always succeeds regardless of
    the original casing the LLM chose for parameter names.
    """
    out: dict[str, Any] = {}
    for k, v in d.items():
        lk = k.lower()
        if lk not in out:
            out[lk] = v
    return out


def score_row(
    expected_tool: str,
    expected_params: dict[str, Any],
    predicted_tool: str,
    predicted_params: dict[str, Any],
) -> dict[str, float]:
    gt_params = _lower_dict(expected_params or {})
    pred_params = _lower_dict(predicted_params or {})

    # FSP
    fsp = 1.0 if str(predicted_tool).strip().lower() == str(expected_tool).strip().lower() else 0.0

    # ACS — F1 on argument name overlap.
    # Exclude the positional 'df' argument from predicted keys; it is always
    # present in parsed output but never in the ground-truth expected dict.
    gt_arg_names = set(gt_params.keys())
    pred_arg_names = {k for k in pred_params.keys() if k != "df"}

    # FIX: when both sides have zero params, the match is perfect → ACS = 1.0.
    if not gt_arg_names and not pred_arg_names:
        acs = 1.0
    else:
        tp = len(gt_arg_names & pred_arg_names)
        fp = len(pred_arg_names - gt_arg_names)
        fn = len(gt_arg_names - pred_arg_names)
        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        acs = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0

    # AVC — now uses lowered dicts so `.get(key)` is safe.
    if gt_params:
        n_correct = 0
        for key in gt_arg_names & pred_arg_names:
            gt_val = str(gt_params.get(key, "")).strip().lower()
            pred_val = str(pred_params.get(key, "")).strip().lower()
            if gt_val and pred_val and gt_val == pred_val:
                n_correct += 1
        avc = n_correct / max(len(gt_arg_names & pred_arg_names), 1)
    else:
        avc = 1.0 if predicted_tool == expected_tool else 0.0

    # Task Acc
    gt_set = expected_set(expected_tool, gt_params)
    pred_set = predicted_set(predicted_tool, pred_params)
    tp_t = len(gt_set & pred_set)
    fp_t = len(pred_set - gt_set)
    fn_t = len(gt_set - pred_set)
    precision_t = tp_t / (tp_t + fp_t) if (tp_t + fp_t) else 0.0
    recall_t = tp_t / (tp_t + fn_t) if (tp_t + fn_t) else 0.0
    task_acc = (2 * precision_t * recall_t / (precision_t + recall_t)) if (precision_t + recall_t) else 0.0
    correct_ratio = 1.0 if task_acc >= 1.0 else 0.0

    return {
        "FSP": round(fsp, 4),
        "ACS": round(acs, 4),
        "AVC": round(avc, 4),
        "Task_Acc": round(task_acc, 4),
        "Correct_Ratio": round(correct_ratio, 4),
    }


def make_validation_report(before: dict[str, Any], after: dict[str, Any]) -> str:
    if before == after:
        return ""
    return f"{before} -> {after}"


# -----------------------------------------------------------------------------
# Error taxonomy
# -----------------------------------------------------------------------------
ERROR_TYPES = [
    "CORRECT",
    "TOOL_WRONG",
    "PARSE_FAILURE",
    "PARAM_MISSING",
    "PARAM_VALUE_WRONG",
    "PARAM_EXTRA",
    "FORMAT_VIOLATION",
]


def classify_error(
    expected_tool: str,
    expected_params: dict[str, Any],
    predicted_tool: str,
    predicted_params: dict[str, Any],
    raw_output: str,
) -> str:
    if predicted_tool not in ALLOWED_TOOLS:
        raw_lower = (raw_output or "").lower().strip()
        if any(marker in raw_lower for marker in ["```", "i think", "based on", "the answer"]):
            return "FORMAT_VIOLATION"
        return "PARSE_FAILURE"

    if str(predicted_tool).strip().lower() != str(expected_tool).strip().lower():
        return "TOOL_WRONG"

    # FIX: use lowered dicts for case-insensitive key access (same as score_row).
    gt = _lower_dict(expected_params or {})
    pred = _lower_dict(predicted_params or {})
    gt_keys = set(gt.keys())
    pred_keys = {k for k in pred.keys() if k != "df"}

    if not gt_keys:
        return "CORRECT"

    missing = gt_keys - pred_keys
    if missing:
        return "PARAM_MISSING"

    for key in gt_keys & pred_keys:
        gt_val = str(gt.get(key, "")).strip().lower()
        pred_val = str(pred.get(key, "")).strip().lower()
        if gt_val and pred_val and gt_val != pred_val:
            return "PARAM_VALUE_WRONG"

    extra = pred_keys - gt_keys
    if extra:
        return "PARAM_EXTRA"

    return "CORRECT"


# -----------------------------------------------------------------------------
# Bootstrap confidence intervals
# -----------------------------------------------------------------------------
def bootstrap_ci(
    values: np.ndarray,
    stat_fn=np.mean,
    n_boot: int = 2000,
    ci: float = 0.95,
    seed: int = 42,
) -> tuple[float, float, float]:
    rng = np.random.default_rng(seed)
    values = np.asarray(values, dtype=float)
    if len(values) == 0:
        return 0.0, 0.0, 0.0
    point = float(stat_fn(values))
    boot_stats = np.array([
        stat_fn(rng.choice(values, size=len(values), replace=True))
        for _ in range(n_boot)
    ])
    alpha = (1 - ci) / 2
    lo = float(np.percentile(boot_stats, alpha * 100))
    hi = float(np.percentile(boot_stats, (1 - alpha) * 100))
    return point, lo, hi


# -----------------------------------------------------------------------------
# Inference wrapper
# -----------------------------------------------------------------------------
def run_llm(
    model_id: str,
    messages: list[dict[str, str]],
    temperature: Optional[float] = None,
) -> tuple[str, float, Optional[str]]:
    t0 = time.perf_counter()
    if temperature is None:
        temperature = TEMPERATURE
    try:
        text, elapsed, err = generate(
            backend="llamacpp",
            model_id=model_id,
            messages=messages,
            temperature=temperature,
            enable_thinking=False,
        )
        if err:
            return text or "", elapsed, err
        return text or "", elapsed, None
    except Exception as e:
        return "", time.perf_counter() - t0, str(e)


# -----------------------------------------------------------------------------
# Benchmarking
# -----------------------------------------------------------------------------
@dataclass
class RunConfig:
    mode: str
    reps: int
    models: list[str]
    output_dir: Path
    output_basename: str
    warmup_reps: int = 1


def build_model_list(models_arg: Optional[str]) -> list[str]:
    available = get_llamacpp_models()
    if models_arg:
        requested = [m.strip() for m in models_arg.split(",") if m.strip()]
        if requested:
            return requested
    return available


def normalize_model_name(model_id: str) -> str:
    base = model_id.split("/")[-1]
    return base[:-5] if base.lower().endswith(".gguf") else base


def open_checkpoint_csv(path: Path, resume: bool = False):
    """
    Open a CSV checkpoint file for writing.

    Args:
        path:   Destination file path.
        resume: If True and the file already exists, append rows (skip header).
                If False (default), overwrite any existing file — prevents
                duplicate rows from re-runs with the same --basename.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    # FIX: default to overwrite ("w") so re-runs don't silently accumulate
    # duplicate rows. Pass resume=True (via --resume flag) to append instead.
    if resume and path.exists():
        mode = "a"
        write_header = False
    else:
        mode = "w"
        write_header = True
    fh = path.open(mode, newline="", encoding="utf-8")
    writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
    if write_header:
        writer.writeheader()
        fh.flush()
    return writer, fh


CSV_FIELDS = [
    "Run_ID", "Run_Timestamp", "Benchmark_Mode", "Validation_Mode",
    "Model", "Model_Full",
    "TC_ID", "TC_Group", "TC_Difficulty", "Question",
    "Expected_Tool", "Expected_Params_JSON",
    "Rep",
    "Router_Time_s", "Specialist_Time_s", "Agent_Time_s", "Total_Time_s",
    "Peak_CPU_Pct", "Peak_RAM_MB",
    "Prompt_Tokens", "Completion_Tokens",
    "Router_Raw", "Specialist_Raw",
    "Pred_Tool_Raw", "Pred_Params_Raw_JSON",
    "Pred_Tool_Validated", "Pred_Params_Validated_JSON",
    "Validation_Warnings",
    "FSP_Raw", "ACS_Raw", "AVC_Raw", "Task_Acc_Raw", "Correct_Ratio_Raw",
    "FSP_Validated", "ACS_Validated", "AVC_Validated", "Task_Acc_Validated", "Correct_Ratio_Validated",
    "Error_Type_Raw", "Error_Type_Validated",
    "Raw_Parse_OK", "Validated_Parse_OK",
    "Notes",
]


def format_expected_params(expected: dict[str, Any]) -> str:
    return json.dumps(expected, ensure_ascii=False, sort_keys=True)


def benchmark_single_agent(
    model_id: str,
    tc: dict[str, Any],
    rep: int,
    current_date: str,
    metadata: str,
) -> dict[str, Any]:
    prompt = build_single_agent_prompt(metadata=metadata, current_date=current_date)
    with ResourceMonitor() as rm:
        raw_text, elapsed, err = run_llm(
            model_id,
            messages=[
                {"role": "system", "content": prompt},
                {"role": "user", "content": tc["q"]},
            ],
        )
    # FIX: capture token usage immediately after the LLM call while it's still
    # the "last" call — avoids stale values from earlier calls in the process.
    usage = get_last_usage()
    tool, kwargs, cleaned = extract_first_tool_call(raw_text)
    kwargs = canonicalize_params(kwargs)
    validated_kwargs, warning = validate_and_fix_params(kwargs, VALIDATION_DF)
    validated_kwargs = canonicalize_params(validated_kwargs)
    return {
        "router_time_s": 0.0,
        "specialist_time_s": elapsed,
        "agent_time_s": elapsed,
        "total_time_s": elapsed,
        "peak_cpu_pct": rm.peak_cpu_percent,
        "peak_ram_mb": rm.peak_ram_mb,
        "router_raw": tool,
        "specialist_raw": cleaned,
        "pred_tool_raw": tool,
        "pred_params_raw": kwargs,
        "pred_tool_validated": tool,
        "pred_params_validated": validated_kwargs,
        "warning": warning or "",
        "raw_parse_ok": 1 if tool in ALLOWED_TOOLS else 0,
        "validated_parse_ok": 1 if tool in ALLOWED_TOOLS else 0,
        "notes": err or "",
        "_prompt_tokens": usage.get("prompt_tokens", 0),
        "_completion_tokens": usage.get("completion_tokens", 0),
    }


def benchmark_dual_agent(
    model_id: str,
    tc: dict[str, Any],
    rep: int,
    current_date: str,
    metadata: str,
    router_model_id: Optional[str] = None,
) -> dict[str, Any]:
    router_model = router_model_id or model_id
    specialist_model = model_id

    with ResourceMonitor() as rm:
        # Stage 1: Router
        router_raw, router_time, router_err = run_llm(
            router_model,
            messages=[
                {"role": "system", "content": ROUTER_PROMPT},
                {"role": "user", "content": tc["q"]},
            ],
        )
        # FIX: capture router token usage before the specialist call overwrites it.
        router_usage = get_last_usage()

        router_tool, _, _ = extract_first_tool_call(router_raw)
        router_tool = router_tool.strip().strip("`'\"")
        if router_tool not in ALLOWED_TOOLS:
            router_tool = router_raw.split()[0].strip().strip("`'\"")

        # Stage 2: Specialist
        predicted_tool_for_prompt = router_tool if router_tool in ALLOWED_TOOLS else "calculate_total"
        tool_prompt_template = get_tool_prompt(predicted_tool_for_prompt)
        if tool_prompt_template is None:
            tool_prompt_template = get_tool_prompt("calculate_total")

        specialist_system_prompt = tool_prompt_template.format(
            metadata=metadata, current_date=current_date
        )

        specialist_raw, specialist_time, spec_err = run_llm(
            specialist_model,
            messages=[
                {"role": "system", "content": specialist_system_prompt},
                {"role": "user", "content": tc["q"]},
            ],
        )
        # FIX: capture specialist usage before exiting the monitor context.
        specialist_usage = get_last_usage()

    spec_tool, spec_kwargs, cleaned = extract_first_tool_call(specialist_raw)
    spec_kwargs = canonicalize_params(spec_kwargs)
    validated_kwargs, warning = validate_and_fix_params(spec_kwargs, VALIDATION_DF)
    validated_kwargs = canonicalize_params(validated_kwargs)

    return {
        "router_time_s": router_time,
        "specialist_time_s": specialist_time,
        "agent_time_s": 0.0,
        "total_time_s": router_time + specialist_time,
        "peak_cpu_pct": rm.peak_cpu_percent,
        "peak_ram_mb": rm.peak_ram_mb,
        "router_raw": router_raw,
        "specialist_raw": cleaned,
        "pred_tool_raw": router_tool,
        "pred_params_raw": spec_kwargs,
        "pred_tool_validated": router_tool,
        "pred_params_validated": validated_kwargs,
        "warning": warning or "",
        "raw_parse_ok": 1 if router_tool in ALLOWED_TOOLS and spec_tool in ALLOWED_TOOLS else 0,
        "validated_parse_ok": 1 if router_tool in ALLOWED_TOOLS and spec_tool in ALLOWED_TOOLS else 0,
        "notes": "; ".join([x for x in [router_err, spec_err] if x]) if (router_err or spec_err) else "",
        # FIX: sum router + specialist token counts so dual-mode reports the full
        # prompt budget consumed, not just the specialist stage.
        "_prompt_tokens": (
            router_usage.get("prompt_tokens", 0) + specialist_usage.get("prompt_tokens", 0)
        ),
        "_completion_tokens": (
            router_usage.get("completion_tokens", 0) + specialist_usage.get("completion_tokens", 0)
        ),
    }


def benchmark_model(
    model_id: str,
    cfg: RunConfig,
    ckpt_writer: csv.DictWriter,
    ckpt_fh,
    router_model_id: Optional[str] = None,
) -> list[dict[str, Any]]:
    short = normalize_model_name(model_id)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    current_date = datetime.now().strftime("%Y-%m-%d")

    rows: list[dict[str, Any]] = []

    logger.info("Running model %s (%s) in %s mode", short, model_id, cfg.mode)

    if cfg.warmup_reps > 0:
        try:
            warmup_model(model_id, current_date, METADATA_TEXT, n_reps=cfg.warmup_reps)
        except Exception:
            pass

    for tc in TEST_CASES:
        logger.info("TC %s | %s", tc["id"], tc["q"])
        for rep in range(1, cfg.reps + 1):
            if cfg.mode == "single":
                result = benchmark_single_agent(model_id, tc, rep, current_date, METADATA_TEXT)
                benchmark_mode = "single"
            elif cfg.mode == "dual":
                result = benchmark_dual_agent(model_id, tc, rep, current_date, METADATA_TEXT, router_model_id=router_model_id)
                benchmark_mode = "dual"
            else:
                raise ValueError(f"Unknown mode: {cfg.mode}")

            raw_tool = result["pred_tool_raw"]
            raw_params = result["pred_params_raw"]
            validated_tool = result["pred_tool_validated"]
            validated_params = result["pred_params_validated"]

            raw_scores = score_row(tc["tool"], tc["expected"], raw_tool, raw_params)
            validated_scores = score_row(tc["tool"], tc["expected"], validated_tool, validated_params)

            specialist_raw_text = str(result.get("specialist_raw", ""))
            error_raw = classify_error(tc["tool"], tc["expected"], raw_tool, raw_params, specialist_raw_text)
            error_val = classify_error(tc["tool"], tc["expected"], validated_tool, validated_params, specialist_raw_text)

            # FIX: use token counts stored in result dict rather than calling
            # get_last_usage() here — for dual mode this would only return
            # specialist tokens, undercounting the router's prompt budget.
            prompt_tokens = result.get("_prompt_tokens", 0)
            completion_tokens = result.get("_completion_tokens", 0)

            row = {
                "Run_ID": run_id,
                "Run_Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Benchmark_Mode": benchmark_mode,
                "Validation_Mode": "raw+validated",
                "Model": short,
                "Model_Full": model_id,
                "TC_ID": tc["id"],
                "TC_Group": tc["group"],
                "TC_Difficulty": tc.get("difficulty", "unknown"),
                "Question": tc["q"],
                "Expected_Tool": tc["tool"],
                "Expected_Params_JSON": format_expected_params(tc["expected"]),
                "Rep": rep,
                "Router_Time_s": round(result["router_time_s"], 4),
                "Specialist_Time_s": round(result["specialist_time_s"], 4),
                "Agent_Time_s": round(result["agent_time_s"], 4),
                "Total_Time_s": round(result["total_time_s"], 4),
                "Peak_CPU_Pct": round(result["peak_cpu_pct"], 2),
                "Peak_RAM_MB": round(result["peak_ram_mb"], 2),
                "Prompt_Tokens": prompt_tokens,
                "Completion_Tokens": completion_tokens,
                "Router_Raw": str(result["router_raw"])[:500],
                "Specialist_Raw": specialist_raw_text[:1200],
                "Pred_Tool_Raw": raw_tool,
                "Pred_Params_Raw_JSON": json.dumps(raw_params, ensure_ascii=False, sort_keys=True),
                "Pred_Tool_Validated": validated_tool,
                "Pred_Params_Validated_JSON": json.dumps(validated_params, ensure_ascii=False, sort_keys=True),
                "Validation_Warnings": result["warning"],
                "FSP_Raw": raw_scores["FSP"],
                "ACS_Raw": raw_scores["ACS"],
                "AVC_Raw": raw_scores["AVC"],
                "Task_Acc_Raw": raw_scores["Task_Acc"],
                "Correct_Ratio_Raw": raw_scores["Correct_Ratio"],
                "FSP_Validated": validated_scores["FSP"],
                "ACS_Validated": validated_scores["ACS"],
                "AVC_Validated": validated_scores["AVC"],
                "Task_Acc_Validated": validated_scores["Task_Acc"],
                "Correct_Ratio_Validated": validated_scores["Correct_Ratio"],
                "Error_Type_Raw": error_raw,
                "Error_Type_Validated": error_val,
                "Raw_Parse_OK": result["raw_parse_ok"],
                "Validated_Parse_OK": result["validated_parse_ok"],
                "Notes": result["notes"],
            }
            ckpt_writer.writerow(row)
            ckpt_fh.flush()
            rows.append(row)

            status = "✓" if validated_scores["Task_Acc"] >= 1.0 else ("~" if validated_scores["Task_Acc"] > 0 else "✗")
            logger.info(
                "[%s] %s rep %d | total=%.2fs | cpu=%.1f%% ram=%.0fMB | raw tool=%s | raw task=%.2f | val task=%.2f | err=%s",
                tc["id"], status, rep, result["total_time_s"], result["peak_cpu_pct"], result["peak_ram_mb"],
                raw_tool, raw_scores["Task_Acc"], validated_scores["Task_Acc"], error_raw,
            )

    return rows


# -----------------------------------------------------------------------------
# Excel summary
# -----------------------------------------------------------------------------
def _pct(series: pd.Series) -> float:
    return float(series.mean() * 100.0) if len(series) else 0.0


def build_summary_frames(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    if df.empty:
        return {
            "overall": pd.DataFrame(),
            "by_case": pd.DataFrame(),
            "by_model": pd.DataFrame(),
            "by_group": pd.DataFrame(),
        }

    rows = []
    for (bench_mode, model), grp in df.groupby(["Benchmark_Mode", "Model"]):
        row = {
            "Benchmark_Mode": bench_mode,
            "Model": model,
            "Reps": int(len(grp)),
            "Avg_Total_s": grp["Total_Time_s"].mean(),
            "Median_Total_s": grp["Total_Time_s"].median(),
            "Avg_RAM_MB": grp["Peak_RAM_MB"].mean(),
            "Avg_CPU_Pct": grp["Peak_CPU_Pct"].mean(),
            # FIX: include token budget in model-level summary for paper efficiency table.
            "Avg_Prompt_Tokens": grp["Prompt_Tokens"].mean() if "Prompt_Tokens" in grp.columns else 0,
            "Avg_Completion_Tokens": grp["Completion_Tokens"].mean() if "Completion_Tokens" in grp.columns else 0,
            "Raw_FSP": grp["FSP_Raw"].mean(),
            "Raw_ACS": grp["ACS_Raw"].mean(),
            "Raw_AVC": grp["AVC_Raw"].mean(),
            "Raw_Task_Acc": grp["Task_Acc_Raw"].mean(),
            "Raw_Correct_Ratio": grp["Correct_Ratio_Raw"].mean(),
            "Validated_FSP": grp["FSP_Validated"].mean(),
            "Validated_ACS": grp["ACS_Validated"].mean(),
            "Validated_AVC": grp["AVC_Validated"].mean(),
            "Validated_Task_Acc": grp["Task_Acc_Validated"].mean(),
            "Validated_Correct_Ratio": grp["Correct_Ratio_Validated"].mean(),
            "Delta_Task_Acc": grp["Task_Acc_Validated"].mean() - grp["Task_Acc_Raw"].mean(),
            "Delta_ACS": grp["ACS_Validated"].mean() - grp["ACS_Raw"].mean(),
            "Delta_AVC": grp["AVC_Validated"].mean() - grp["AVC_Raw"].mean(),
        }
        rows.append(row)
    by_model = pd.DataFrame(rows)

    by_case = (
        df.groupby(["Benchmark_Mode", "TC_Group"], as_index=False)
          .agg(
              Raw_Task_Acc=("Task_Acc_Raw", "mean"),
              Validated_Task_Acc=("Task_Acc_Validated", "mean"),
              Raw_FSP=("FSP_Raw", "mean"),
              Validated_FSP=("FSP_Validated", "mean"),
              Raw_ACS=("ACS_Raw", "mean"),
              Validated_ACS=("ACS_Validated", "mean"),
              Avg_Total_s=("Total_Time_s", "mean"),
              Avg_RAM_MB=("Peak_RAM_MB", "mean"),
              Avg_CPU_Pct=("Peak_CPU_Pct", "mean"),
          )
    )

    overall = pd.DataFrame(
        [{
            "Rows": len(df),
            "Models": df["Model"].nunique(),
            "Benchmark_Modes": ", ".join(sorted(df["Benchmark_Mode"].unique())),
            "Avg_Total_s": df["Total_Time_s"].mean(),
            "Avg_RAM_MB": df["Peak_RAM_MB"].mean(),
            "Avg_CPU_Pct": df["Peak_CPU_Pct"].mean(),
            "Raw_FSP": df["FSP_Raw"].mean(),
            "Raw_ACS": df["ACS_Raw"].mean(),
            "Raw_AVC": df["AVC_Raw"].mean(),
            "Raw_Task_Acc": df["Task_Acc_Raw"].mean(),
            "Raw_Correct_Ratio": df["Correct_Ratio_Raw"].mean(),
            "Validated_FSP": df["FSP_Validated"].mean(),
            "Validated_ACS": df["ACS_Validated"].mean(),
            "Validated_AVC": df["AVC_Validated"].mean(),
            "Validated_Task_Acc": df["Task_Acc_Validated"].mean(),
            "Validated_Correct_Ratio": df["Correct_Ratio_Validated"].mean(),
        }]
    )

    by_group = (
        df.groupby(["Benchmark_Mode", "TC_Group"], as_index=False)
          .agg(
              Raw_Task_Acc=("Task_Acc_Raw", "mean"),
              Validated_Task_Acc=("Task_Acc_Validated", "mean"),
              Raw_ACS=("ACS_Raw", "mean"),
              Validated_ACS=("ACS_Validated", "mean"),
              Raw_AVC=("AVC_Raw", "mean"),
              Validated_AVC=("AVC_Validated", "mean"),
              Avg_Total_s=("Total_Time_s", "mean"),
              Avg_RAM_MB=("Peak_RAM_MB", "mean"),
          )
    )
    by_group["Delta_Task_Acc"] = by_group["Validated_Task_Acc"] - by_group["Raw_Task_Acc"]

    # Bootstrap CI on by_model
    ci_rows = []
    for (bench_mode, model), grp in df.groupby(["Benchmark_Mode", "Model"]):
        ta_point, ta_lo, ta_hi = bootstrap_ci(grp["Task_Acc_Validated"].values)
        fsp_point, fsp_lo, fsp_hi = bootstrap_ci(grp["FSP_Validated"].values)
        ci_rows.append({
            "Benchmark_Mode": bench_mode, "Model": model,
            "Task_Acc_Mean": round(ta_point, 4),
            "Task_Acc_CI_Low": round(ta_lo, 4),
            "Task_Acc_CI_High": round(ta_hi, 4),
            "FSP_Mean": round(fsp_point, 4),
            "FSP_CI_Low": round(fsp_lo, 4),
            "FSP_CI_High": round(fsp_hi, 4),
        })
    ci_frame = pd.DataFrame(ci_rows)

    if not by_model.empty and not ci_frame.empty:
        by_model = by_model.merge(ci_frame, on=["Benchmark_Mode", "Model"], how="left")

    # Error distribution
    error_dist = pd.DataFrame()
    if "Error_Type_Validated" in df.columns:
        error_dist = (
            df.groupby(["Benchmark_Mode", "Model", "Error_Type_Validated"])
              .size()
              .reset_index(name="Count")
        )
        error_pivot = error_dist.pivot_table(
            index=["Benchmark_Mode", "Model"],
            columns="Error_Type_Validated",
            values="Count",
            fill_value=0,
        ).reset_index()
    else:
        error_pivot = pd.DataFrame()

    return {
        "overall": overall,
        "by_model": by_model.sort_values(["Benchmark_Mode", "Model"]).reset_index(drop=True),
        "by_case": by_case.sort_values(["Benchmark_Mode", "TC_Group"]).reset_index(drop=True),
        "by_group": by_group.sort_values(["Benchmark_Mode", "TC_Group"]).reset_index(drop=True),
        "error_distribution": error_pivot if not error_pivot.empty else error_dist,
    }


def warmup_model(model_id: str, current_date: str, metadata: str, n_reps: int = 1) -> None:
    """
    Warm up the model by running inference across all tool prompts and the router prompt.

    FIX: original only warmed up calculate_total and the router. On Apple Silicon
    llama.cpp, KV-cache cold-start affects the first generation for each distinct
    system prompt. Rotating through all five tool prompts ensures first-rep latency
    measurements reflect steady-state, not initialization overhead.
    """
    warmup_query = "How much did I spend on food?"
    # Router prompt
    run_llm(model_id, [
        {"role": "system", "content": ROUTER_PROMPT},
        {"role": "user", "content": warmup_query},
    ])
    # All specialist prompts
    for tool_name in sorted(ALLOWED_TOOLS):
        tool_prompt = get_tool_prompt(tool_name)
        if tool_prompt is None:
            continue
        try:
            formatted = tool_prompt.format(metadata=metadata, current_date=current_date)
        except KeyError:
            continue
        for _ in range(n_reps):
            run_llm(model_id, [
                {"role": "system", "content": formatted},
                {"role": "user", "content": warmup_query},
            ])


def write_excel(df: pd.DataFrame, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # Raw sheet
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        # FIX: was Font(bold=True) with default black text on dark navy fill —
        # invisible. White text is required for the dark header background.
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(col_name) + 2, 12), 40)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row[col_name])

    # Summary sheets
    frames = build_summary_frames(df)
    for sheet_name, frame in frames.items():
        ws2 = wb.create_sheet(sheet_name[:31])
        ws2.freeze_panes = "A2"
        ws2.sheet_view.showGridLines = False
        if frame.empty:
            ws2["A1"] = "No data"
            continue
        for c_idx, col_name in enumerate(frame.columns, start=1):
            cell = ws2.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws2.column_dimensions[get_column_letter(c_idx)].width = min(max(len(col_name) + 2, 12), 35)
        for r_idx, (_, row) in enumerate(frame.iterrows(), start=2):
            for c_idx, col_name in enumerate(frame.columns, start=1):
                ws2.cell(row=r_idx, column=c_idx, value=row[col_name])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Benchmark single-agent and dual-agent expense routers"
    )
    parser.add_argument(
        "--mode", choices=["single", "dual", "both"], default="both",
        help="Benchmark mode. 'both'=single+dual",
    )
    parser.add_argument("--reps", type=int, default=5, help="Repetitions per question (default 5 for CI)")
    parser.add_argument("--models", type=str, default=None, help="Comma-separated llama.cpp model ids. Default: all available")
    parser.add_argument("--output-dir", type=str, default="benchmark_outputs", help="Output directory")
    parser.add_argument("--basename", type=str, default=None, help="Base filename prefix")
    parser.add_argument("--router-model", type=str, default=None, help="Optional dedicated router model for dual mode")
    parser.add_argument("--quick", action="store_true", help="Shortcut for 1 rep and 1 model")
    parser.add_argument("--temperature", type=float, default=None, help="Sampling temperature")
    # FIX: added --resume flag. Default behaviour is now to overwrite the checkpoint CSV,
    # preventing silent row duplication on re-runs with the same --basename.
    parser.add_argument(
        "--resume", action="store_true",
        help="Append to an existing CSV checkpoint instead of overwriting it",
    )
    args = parser.parse_args()

    if args.temperature is not None:
        global TEMPERATURE
        TEMPERATURE = args.temperature

    reps = 1 if args.quick else args.reps
    models = build_model_list(args.models)
    if args.quick and models:
        models = models[:1]

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = args.basename or f"expense_router_benchmark_{stamp}"

    if args.mode == "both":
        modes = ["single", "dual"]
    else:
        modes = [args.mode]

    all_rows: list[dict[str, Any]] = []

    logger.info("Backend root: %s", BACKEND_ROOT)
    logger.info("Models: %s", models)
    logger.info("Modes: %s", modes)
    logger.info("Reps: %d", reps)
    if args.resume:
        logger.info("Resume mode: appending to existing checkpoints")

    for mode in modes:
        csv_path = out_dir / f"{base}_{mode}.csv"
        xlsx_path = out_dir / f"{base}_{mode}.xlsx"
        ckpt_writer, ckpt_fh = open_checkpoint_csv(csv_path, resume=args.resume)
        try:
            for model_id in models:
                cfg = RunConfig(
                    mode=mode, reps=reps, models=models,
                    output_dir=out_dir, output_basename=base,
                )
                rows = benchmark_model(model_id, cfg, ckpt_writer, ckpt_fh, router_model_id=args.router_model)
                all_rows.extend(rows)
                if len(models) > 1:
                    try:
                        free_model_memory()
                    except Exception:
                        pass
                gc.collect()
        finally:
            ckpt_fh.close()

        df = pd.read_csv(csv_path)
        write_excel(df, xlsx_path)
        logger.info("Wrote %s and %s", csv_path, xlsx_path)

    if len(modes) > 1:
        combined_csv = out_dir / f"{base}_combined.csv"
        combined_xlsx = out_dir / f"{base}_combined.xlsx"
        combined_df = pd.DataFrame(all_rows)
        combined_df.to_csv(combined_csv, index=False)
        write_excel(combined_df, combined_xlsx)
        logger.info("Wrote combined outputs: %s / %s", combined_csv, combined_xlsx)


if __name__ == "__main__":
    main()